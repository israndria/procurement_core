"""Reader/writer Master Data PPK V2.

Menjaga layout legacy ``Master Data`` dan membaca tabel domain V2 tanpa
menyimpan ulang workbook XLSM melalui openpyxl.
"""

from __future__ import annotations

import copy
import json
import os
import shutil
import tempfile
from pathlib import Path
from typing import Any


MASTER_SHEET = "Master Data"
JKK_SHEET = "Data JKK"
PK_SHEET = "Data PK"
SNAPSHOT_VERSION = 2


def _value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _header(value: Any) -> str:
    """Normalisasi header tampilan sederhana ke key internal stabil."""
    text = str(value or "").strip()
    return "No" if text.rstrip(".").lower() == "no" else text


def _read_workbook(excel_path: str | os.PathLike[str]):
    """Buka salinan XLSM read-only; caller wajib menutup workbook."""
    from openpyxl import load_workbook

    source = Path(excel_path)
    temp_dir = Path(tempfile.mkdtemp(prefix="master_data_v2_"))
    temp_path = temp_dir / source.name
    shutil.copy2(source, temp_path)
    wb = load_workbook(
        temp_path,
        # Excel Table metadata is unavailable on ReadOnlyWorksheet.
        # Workbook is a temporary copy and is never saved.
        read_only=False,
        data_only=True,
        keep_vba=True,
        keep_links=False,
    )
    return wb, temp_dir


def _close_workbook(wb: Any, temp_dir: Path) -> None:
    try:
        wb.close()
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def read_master_data(excel_path: str | os.PathLike[str]) -> dict[str, Any]:
    """Baca label kolom A dan nilai kolom B pada sheet legacy."""
    wb, temp_dir = _read_workbook(excel_path)
    try:
        if MASTER_SHEET not in wb.sheetnames:
            raise ValueError(f"Sheet {MASTER_SHEET!r} tidak ditemukan")
        ws = wb[MASTER_SHEET]
        result: dict[str, Any] = {}
        for row in ws.iter_rows(min_col=1, max_col=2):
            label = row[0].value
            if label is None:
                continue
            result[str(label).strip()] = _value(row[1].value)
        return result
    finally:
        _close_workbook(wb, temp_dir)


def _read_table(ws: Any, table_name: str) -> list[dict[str, Any]]:
    """Baca Excel Table berdasarkan nama stabil, tanpa bergantung posisi kolom."""
    table = ws.tables.get(table_name)
    if table is None:
        return []
    cells = ws[table.ref]
    if not cells:
        return []
    headers = [_header(cell.value) for cell in cells[0]]
    rows: list[dict[str, Any]] = []
    for row in cells[1:]:
        item = {header: _value(cell.value) for header, cell in zip(headers, row) if header}
        # Nomor urut template terisi otomatis; jangan anggap baris kosong
        # sebagai record hanya karena kolom No berisi angka.
        payload = [value for key, value in item.items() if key.lower() not in {"no", "aktif"}]
        if any(str(value).strip() for value in payload):
            rows.append(item)
    return rows


def read_domain_data(excel_path: str | os.PathLike[str]) -> dict[str, list[dict[str, Any]]]:
    """Baca tabel domain JKK/PK. Hanya data aktif yang dikembalikan."""
    wb, temp_dir = _read_workbook(excel_path)
    try:
        # V2 terbaru menyatukan tabel JKK ke Data PK. Data JKK lama tetap
        # dibaca sebagai fallback compatibility.
        if PK_SHEET in wb.sheetnames and "tblJKKPersonil" in wb[PK_SHEET].tables:
            jkk = _read_table(wb[PK_SHEET], "tblJKKPersonil")
        elif JKK_SHEET in wb.sheetnames and "tblJKKPersonil_Legacy" in wb[JKK_SHEET].tables:
            jkk = _read_table(wb[JKK_SHEET], "tblJKKPersonil_Legacy")
        elif JKK_SHEET in wb.sheetnames:
            jkk = _read_table(wb[JKK_SHEET], "tblJKKPersonil")
        else:
            jkk = []
        pk_personil = _read_table(wb[PK_SHEET], "tblPKPersonil") if PK_SHEET in wb.sheetnames else []
        pk_alat = _read_table(wb[PK_SHEET], "tblPKPeralatan") if PK_SHEET in wb.sheetnames else []

        def active(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
            return [row for row in rows if str(row.get("Aktif", "Ya")).strip().lower() not in {"tidak", "no", "0"}]

        return {
            "jkk_personil": active(jkk),
            "pk_personil": active(pk_personil),
            "pk_peralatan": active(pk_alat),
        }
    finally:
        _close_workbook(wb, temp_dir)


def read_master_data_v2(excel_path: str | os.PathLike[str]) -> dict[str, Any]:
    return {"master_data": read_master_data(excel_path), "domain": read_domain_data(excel_path)}


def build_snapshot(excel_path: str | os.PathLike[str]) -> dict[str, Any]:
    """Bangun snapshot V2 sambil mempertahankan field flat legacy."""
    common = read_master_data(excel_path)
    domain = read_domain_data(excel_path)
    snapshot = dict(common)
    snapshot.update(
        {
            "_schema_version": SNAPSHOT_VERSION,
            "master_data": copy.deepcopy(common),
            "data_jkk": {"personil": domain["jkk_personil"]},
            "data_pk": {
                "personil": domain["pk_personil"],
                "peralatan": domain["pk_peralatan"],
            },
        }
    )
    return snapshot


def normalize_snapshot(snapshot: dict[str, Any] | None) -> dict[str, Any]:
    """Normalisasi snapshot lama flat atau snapshot V2 ke struktur V2."""
    raw = snapshot or {}
    common = raw.get("master_data")
    if not isinstance(common, dict):
        common = {k: v for k, v in raw.items() if not k.startswith("_") and k not in {"data_jkk", "data_pk"}}
    jkk = raw.get("data_jkk") if isinstance(raw.get("data_jkk"), dict) else {}
    pk = raw.get("data_pk") if isinstance(raw.get("data_pk"), dict) else {}
    return {
        "_schema_version": int(raw.get("_schema_version", 1) or 1),
        "master_data": common,
        "data_jkk": {"personil": jkk.get("personil", []) if isinstance(jkk, dict) else []},
        "data_pk": {
            "personil": pk.get("personil", []) if isinstance(pk, dict) else [],
            "peralatan": pk.get("peralatan", []) if isinstance(pk, dict) else [],
        },
    }


def write_snapshot_json(excel_path: str | os.PathLike[str], output_path: str | os.PathLike[str]) -> None:
    Path(output_path).write_text(json.dumps(build_snapshot(excel_path), ensure_ascii=False, indent=2), encoding="utf-8")


def sync_daftar_paket_snapshot(excel_path: str | os.PathLike[str]) -> None:
    """Update kolom Snapshot pada Daftar Paket via Excel COM.

    Baris lama dicari memakai Kode RUP; bila belum ada, dibuat satu baris baru.
    Snapshot tetap memuat field flat legacy dan blok V2.
    """
    import pythoncom
    import win32com.client

    common = read_master_data(excel_path)
    snapshot = build_snapshot(excel_path)
    payload = json.dumps(snapshot, ensure_ascii=False, separators=(",", ":"))
    pythoncom.CoInitialize()
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.EnableEvents = False
    wb = None
    try:
        wb = xl.Workbooks.Open(str(Path(excel_path)), UpdateLinks=0, ReadOnly=False)
        ws = wb.Worksheets("Daftar Paket")
        headers = {str(ws.Cells(1, col).Value or "").strip(): col for col in range(1, 12)}
        rup_col = headers.get("Kode RUP", 4)
        snap_col = headers.get("Snapshot", 11)
        target = None
        rup = str(common.get("Kode RUP", "")).strip()
        def key(value: Any) -> str:
            text = str(value or "").strip()
            try:
                number = float(text)
                return str(int(number)) if number.is_integer() else text
            except (TypeError, ValueError):
                return text
        rup_key = key(rup)
        for row in range(2, ws.UsedRange.Rows.Count + ws.UsedRange.Row):
            if key(ws.Cells(row, rup_col).Value) == rup_key and rup_key:
                target = row
                break
        if target is None:
            target = ws.UsedRange.Rows.Count + ws.UsedRange.Row
            if target < 2:
                target = 2
            ws.Cells(target, 1).Value = target - 1
        values = {
            2: common.get("Nama Paket (Singkat)", ""),
            3: common.get("Nama Paket (Lengkap)", ""),
            4: common.get("Kode RUP", ""),
            5: common.get("Pagu Anggaran (Angka)", ""),
            6: common.get("Nilai HPS (Angka)", ""),
            7: common.get("Lokasi Pekerjaan", ""),
            8: common.get("Tanggal KAK & HPS", ""),
            9: common.get("Nama Penyedia", ""),
            snap_col: payload,
        }
        for col, value in values.items():
            ws.Cells(target, col).Value = value
        wb.Save()
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        xl.Quit()
        pythoncom.CoUninitialize()


def restore_snapshot(excel_path: str | os.PathLike[str], snapshot: dict[str, Any]) -> None:
    """Restore snapshot ke XLSM memakai Excel COM; tidak memakai openpyxl.save()."""
    import pythoncom
    import win32com.client

    normalized = normalize_snapshot(snapshot)
    pythoncom.CoInitialize()
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.EnableEvents = False
    wb = None
    try:
        wb = xl.Workbooks.Open(str(Path(excel_path)), UpdateLinks=0, ReadOnly=False)
        ws = wb.Worksheets(MASTER_SHEET)
        labels = {}
        for row in range(1, ws.UsedRange.Rows.Count + ws.UsedRange.Row):
            label = str(ws.Cells(row, 1).Value or "").strip()
            if label:
                labels[label] = row
        for label, value in normalized["master_data"].items():
            if label in labels:
                ws.Cells(labels[label], 2).Value = value

        try:
            jkk_ws = _com_sheet_with_table(wb, "tblJKKPersonil", [PK_SHEET])
            jkk_name = "tblJKKPersonil"
        except ValueError:
            jkk_ws = _com_sheet_with_table(wb, "tblJKKPersonil_Legacy", [JKK_SHEET])
            jkk_name = "tblJKKPersonil_Legacy"
        _restore_table(jkk_ws, jkk_name, normalized["data_jkk"]["personil"])
        _restore_table(wb.Worksheets(PK_SHEET), "tblPKPersonil", normalized["data_pk"]["personil"])
        _restore_table(wb.Worksheets(PK_SHEET), "tblPKPeralatan", normalized["data_pk"]["peralatan"])
        wb.Save()
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        xl.Quit()
        pythoncom.CoUninitialize()


def _restore_table(ws: Any, table_name: str, rows: list[dict[str, Any]]) -> None:
    table = ws.ListObjects(table_name)
    headers = [_header(table.HeaderRowRange.Cells(1, col).Value) for col in range(1, table.ListColumns.Count + 1)]
    first_row = table.DataBodyRange.Row if table.ListRows.Count else table.HeaderRowRange.Row + 1
    for offset in range(max(table.ListRows.Count, len(rows))):
        excel_row = first_row + offset
        values = rows[offset] if offset < len(rows) else {}
        for col, header in enumerate(headers, 1):
            ws.Cells(excel_row, col).Value = values.get(header, "")


def _com_sheet_with_table(wb: Any, table_name: str, preferred: list[str]) -> Any:
    """Cari sheet pemilik ListObject; preferred pertama menjadi source of truth."""
    for sheet_name in preferred:
        try:
            ws = wb.Worksheets(sheet_name)
            ws.ListObjects(table_name)
            return ws
        except Exception:
            continue
    raise ValueError(f"Tabel {table_name!r} tidak ditemukan")
