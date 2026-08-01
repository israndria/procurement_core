"""Validasi read-only baseline tampilan Master Data PPK V2.

Validator tidak menyimpan workbook dan tidak melakukan repair otomatis.
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_TABLES = {
    "tblPKPersonil": "A4:D19",
    "tblPKPeralatan": "F4:I19",
    "tblJKKPersonil": "K4:N19",
}
EXPECTED_HEADERS = {
    "tblPKPersonil": ["No.", "Jabatan", "Sertifikat", "Pengalaman Kerja"],
    "tblPKPeralatan": ["No.", "Jenis Alat", "Kapasitas (Minimal)", "Jumlah"],
    "tblJKKPersonil": ["No.", "Jabatan", "Sertifikat", "Pengalaman Kerja"],
}
EXPECTED_HEADER_COLOR_INDEX = {"A4": 56, "F4": 3, "K4": 10}


def _result() -> dict[str, Any]:
    return {"ok": True, "errors": [], "warnings": [], "checks": {}}


def _error(result: dict[str, Any], message: str) -> None:
    result["ok"] = False
    result["errors"].append(message)


def _warning(result: dict[str, Any], message: str) -> None:
    result["warnings"].append(message)


def _check_excel_structure(path: Path, result: dict[str, Any]) -> None:
    try:
        wb = load_workbook(path, read_only=False, data_only=False, keep_vba=True)
    except Exception as exc:
        _error(result, f"Workbook tidak dapat dibaca: {exc}")
        return
    try:
        result["checks"]["sheets"] = wb.sheetnames
        for sheet in ("Data PK", "Master Data", "Data JKK"):
            if sheet not in wb.sheetnames:
                _error(result, f"Sheet wajib tidak ditemukan: {sheet}")
        if "Data PK" not in wb.sheetnames:
            return

        ws = wb["Data PK"]
        merged = {str(item) for item in ws.merged_cells.ranges}
        result["checks"]["merged_data_pk"] = sorted(merged)
        for expected in ("A1:N1", "A2:N2"):
            if expected not in merged:
                _error(result, f"Merged cell Data PK hilang: {expected}")

        table_checks = {}
        for name, expected_ref in EXPECTED_TABLES.items():
            if name not in ws.tables:
                _error(result, f"Tabel Data PK tidak ditemukan: {name}")
                continue
            table = ws.tables[name]
            actual_ref = table.ref
            table_checks[name] = {"ref": actual_ref, "style": table.tableStyleInfo.name if table.tableStyleInfo else None}
            if actual_ref != expected_ref:
                _error(result, f"Range {name} berubah: {actual_ref}; expected {expected_ref}")
            start, end = expected_ref.split(":")
            start_col = ws[start].column
            end_col = ws[end].column
            headers = [ws.cell(ws[start].row, col).value for col in range(start_col, end_col + 1)]
            if headers != EXPECTED_HEADERS[name]:
                _error(result, f"Header {name} berubah: {headers!r}")
        result["checks"]["tables"] = table_checks

        defaults = {"A5": 1, "F5": 1, "K5": 1, "A19": 15, "F19": 15, "K19": 15}
        for coord, expected in defaults.items():
            if ws[coord].value != expected:
                _error(result, f"Default nomor {coord} berubah: {ws[coord].value!r}; expected {expected}")

        master = wb["Master Data"]
        expected_master_labels = {
            31: "Tahap Dokumen",
            72: "Jabatan Direktur",
            93: "Ada Wakil Sah?",
            94: "Nama Wakil Sah",
            95: "Jabatan Wakil Sah",
            96: "Nomor SK Wakil Sah",
            97: "Tanggal SK Wakil Sah",
            101: "NIP Wakil Sah",
        }
        for row, expected in expected_master_labels.items():
            actual = master.cell(row, 1).value
            if actual != expected:
                _error(result, f"Label Master Data A{row} berubah: {actual!r}; expected {expected!r}")
    finally:
        wb.close()


def _check_excel_visual_com(path: Path, result: dict[str, Any]) -> None:
    """Cek warna/font/merge aktual Excel; skip jika COM tidak tersedia."""
    try:
        import pythoncom
        import win32com.client
    except ImportError:
        _warning(result, "pywin32 tidak tersedia; validasi visual COM dilewati")
        return

    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=True)
        ws = wb.Worksheets("Data PK")
        visual = {}
        for coord, expected in EXPECTED_HEADER_COLOR_INDEX.items():
            actual = int(ws.Range(coord).Interior.ColorIndex)
            visual[coord] = actual
            if actual != expected:
                _warning(result, f"Warna header {coord} berubah: {actual}; baseline ColorIndex {expected}")
        for coord, expected_size, expected_bold in (("A1", 14, True), ("A2", 11, False)):
            cell = ws.Range(coord)
            if int(cell.Font.Size) != expected_size or bool(cell.Font.Bold) != expected_bold:
                _warning(result, f"Font {coord} berubah: size={cell.Font.Size}, bold={cell.Font.Bold}")
        result["checks"]["visual_com"] = visual
    except Exception as exc:
        _warning(result, f"Validasi visual COM dilewati: {exc}")
    finally:
        if wb is not None:
            wb.Close(False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def validate(path: str | Path) -> dict[str, Any]:
    target = Path(path)
    result = _result()
    if not target.is_file():
        _error(result, f"File tidak ditemukan: {target}")
        return result
    try:
        with zipfile.ZipFile(target) as archive:
            result["checks"]["zip_test"] = archive.testzip()
            if archive.testzip() is not None:
                _error(result, "ZIP internal workbook rusak")
            if "xl/vbaProject.bin" not in archive.namelist():
                _error(result, "xl/vbaProject.bin tidak ditemukan")
            else:
                result["checks"]["vba_size"] = archive.getinfo("xl/vbaProject.bin").file_size
    except zipfile.BadZipFile as exc:
        _error(result, f"File bukan ZIP XLSM valid: {exc}")
        return result
    _check_excel_structure(target, result)
    _check_excel_visual_com(target, result)
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Path ke Master_Data_PL_PPK.xlsm")
    parser.add_argument("--json", action="store_true", help="Output JSON")
    args = parser.parse_args(argv)
    result = validate(args.workbook)
    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print("OK" if result["ok"] else "FAIL")
        for message in result["errors"]:
            print(f"ERROR: {message}")
        for message in result["warnings"]:
            print(f"WARN: {message}")
    return 0 if result["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
