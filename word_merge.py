"""
word_merge.py - Merge Word template dengan Excel data
=====================================================
Dipanggil dari VBA Excel via Shell (proses terpisah, Excel tidak hang).

Strategy:
  1. Baca data dari Excel yang sedang terbuka (GetObject - cepat)
  2. Copy template ke file (Merged).docx
  3. Buka copy, ganti semua MERGEFIELD dengan data
  4. Tampilkan Word / print / export PDF

Usage:
  python word_merge.py buka    <word_path> <excel_path> <sheet_name>
  python word_merge.py print   <word_path> <excel_path> <sheet_name>
  python word_merge.py pdf     <word_path> <excel_path> <sheet_name> <pdf_name>
  python word_merge.py printer <word_path> <excel_path> <sheet_name> <printer_name> [from_page] [to_page]
"""
import sys
import os
import time
import re
import datetime
import shutil
import glob

from word_xml_compat import normalize_word_document_xml_in_zip


def _safe_filename(s: str, max_len: int = 80) -> str:
    s = re.sub(r'[<>:"/\\|?*]', '', str(s)).strip().replace('\n',' ').replace('\r','')
    return s[:max_len] if s else 'Dokumen'


def _pdf_output_suffix(pdf_name: str = "", package_name: str = "") -> str:
    """Resolve suffix PDF; marker POKJA_NNN mengalahkan nama paket."""
    safe_name = pdf_name or "000"
    marker = "POKJA_"
    if safe_name.upper().startswith(marker):
        return _safe_filename(safe_name[len(marker):])
    return _safe_filename(package_name) if package_name else safe_name


def format_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    s = str(value)
    # Bersihkan line break Excel (\r \n \r\n + literal _x000D_) → spasi tunggal.
    # Excel simpan multiline sebagai CR → muncul "_x000D_" mentah di Word saat merge.
    s = s.replace("_x000D_", "\n").replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"\s*\n\s*", " ", s).strip()
    return s


def normalize_field_name(name):
    s = str(name).strip()
    s = s.replace(" ", "_")
    s = re.sub(r"[^A-Za-z0-9_]", "", s)
    return s


def _parse_currency_text(value):
    """Parse nilai rupiah hasil formula TEXT Excel menjadi float."""
    if value is None:
        return None
    raw = str(value).strip()
    if not raw or raw in {"-", "0"}:
        return None
    raw = re.sub(r"[^0-9,.-]", "", raw)
    if not raw:
        return None
    try:
        return float(raw.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def _normalize_revaluasi_data(data):
    """Samakan field Revaluasi dengan harga final yang dipakai ringkasan."""
    # Halaman awal memakai Harga_Penawaran/Harga_Terkoreksi dari BA klarifikasi
    # (harga final), sedangkan field H1/HPT1 lama menunjuk 4. Evaluasi Harga
    # (harga awal). Gunakan satu sumber final agar satu PDF tidak kontradiktif.
    final_offer = data.get("Harga_Penawaran", "")
    final_corrected = data.get("Harga_Terkoreksi", "")
    if final_offer:
        data["H1"] = final_offer
    if final_corrected:
        data["HPT1"] = final_corrected

    hps = _parse_currency_text(data.get("HPS"))
    corrected = _parse_currency_text(final_corrected)
    if hps and corrected is not None:
        ratio = corrected / hps
        data["persentase_HPS1"] = f"{ratio * 100:.2f}%".replace(".", ",")
        wajar = 0.80 <= ratio <= 1.00
        data["Kewajaran_Harga"] = "WAJAR" if wajar else "TIDAK WAJAR"
        data["Kesimpulan"] = "MEMENUHI" if wajar else "TIDAK MEMENUHI"

    # Bersihkan artefak data lama dari parser/workbook.
    for key in ("Eva_K_34", "Eva_K_38"):
        if data.get(key):
            data[key] = re.sub(r"[.]+$", "", str(data[key]).strip())
    if data.get("Eva_K_47"):
        data["Eva_K_47"] = re.sub(r"\s*\(\s*\)\s*$", "", str(data["Eva_K_47"]).strip())


def read_excel_data(excel_path, sheet_name):
    """Baca data Excel via openpyxl (copy ke temp dulu karena file mungkin terkunci oleh Excel)."""
    import tempfile
    from openpyxl import load_workbook

    data = {}
    temp_dir = tempfile.mkdtemp()
    temp_path = os.path.join(temp_dir, os.path.basename(excel_path))

    try:
        shutil.copy2(excel_path, temp_path)
        wb = load_workbook(temp_path, read_only=True, data_only=True, keep_links=False)

        if sheet_name not in wb.sheetnames:
            wb.close()
            show_error(f"Sheet '{sheet_name}' tidak ditemukan di Excel.")
            return None

        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        values = [c.value for c in ws[2]]
        wb.close()

        # Word mail-merge menomori kolom dengan nama duplikat: occurrence ke-2 dst
        # dapat suffix "1", "2", ... (mis. Hari, Hari1, Hari2). Replikasi agar
        # MERGEFIELD seperti "Harga_Penawaran1"/"Hari1" ketemu saat re-merge.
        seen = {}
        for header, value in zip(headers, values):
            if header:
                header = str(header).strip()
                normalized = normalize_field_name(header)
                val = format_value(value)
                # Formula Excel untuk slot peserta yang tidak terisi sering
                # menghasilkan angka 0. Untuk Word, slot kosong harus benar-
                # benar kosong agar baris peserta tidak tampil sebagai "0".
                if (normalized.startswith("Peserta_") or normalized.startswith("Alamat_")) and val == "0":
                    val = ""
                # Formula TRANSPOSE di sheet KK Evaluasi mengubah sel pemilik
                # kosong menjadi 0 pada sheet `satu_data`. Nilai itu bukan
                # data pemilik yang sah dan tidak boleh masuk ke Revaluasi.
                # Filter berdasarkan field spesifik agar angka 0 yang valid
                # pada evaluasi lain tetap dipertahankan.
                if normalized in {"Eva_K_43", "Eva_K_44", "Eva_K_45"} and val == "0":
                    val = ""
                # nomori per nama-ternormalisasi (sesuai perilaku Word data source):
                # occurrence pertama = nama polos, ke-2 dst = suffix 1,2,...
                # pakai setdefault agar occurrence PERTAMA menang (match Word base).
                n = seen.get(normalized, 0)
                if n == 0:
                    data.setdefault(header, val)
                    if normalized != header:
                        data.setdefault(normalized, val)
                else:
                    data.setdefault(f"{normalized}{n}", val)
                    if header != normalized:
                        data.setdefault(f"{header}{n}", val)
                seen[normalized] = n + 1
        if sheet_name in {"list_dokpil", "list_reviu"}:
            # Tabel persyaratan peralatan pada Dokpil dan Isi Reviu tidak
            # mengambil data dari cache sheet mail-merge. Sumber tunggalnya
            # adalah tabel input yang memang disediakan untuk operator paket.
            data["_source_sheet"] = sheet_name
            data["_dokpil_equipment"] = _read_dokpil_equipment(temp_path)
            data["_dokpil_personnel"] = _read_dokpil_personnel(temp_path)
        if sheet_name == "satu_data":
            _augment_ba_counts(data, temp_path)
            _normalize_revaluasi_data(data)
    except Exception as e:
        show_error(f"Error baca Excel:\n{e}")
        return None
    finally:
        try:
            os.remove(temp_path)
            os.rmdir(temp_dir)
        except:
            pass

    return data


def _meaningful_dokpil_value(value):
    """Normalisasi nilai slot alat; 0 dari formula Excel berarti kosong."""
    value = format_value(value).strip()
    if value.casefold() in ("", "0", "0.0", "none", "null", "-", "—"):
        return ""
    return value


def _dokpil_input_sheet(wb):
    """Cari sheet input alat/personil dengan nama lama maupun nama Tender."""
    aliases = {"tabel alat & personil", "alat & personil"}
    for name in wb.sheetnames:
        if str(name).strip().casefold() in aliases:
            return wb[name]
    return None


def _dokpil_personnel_sheet(wb):
    """Cari sheet personil terpisah, lalu fallback ke sheet gabungan."""
    for name in wb.sheetnames:
        if str(name).strip().casefold() == "personil":
            return wb[name], True
    return _dokpil_input_sheet(wb), False


def _read_dokpil_equipment(excel_copy_path):
    """Baca alat aktif dari sheet input ``B4:E9``.

    Sheet dapat bernama ``Alat & Personil`` (Tender) atau
    ``Tabel Alat & Personil`` (PLPK lama). Baris dengan Jenis Alat
    kosong/0 tidak dikirim ke Dokpil, sehingga jumlah row Word selalu sama
    dengan jumlah alat aktif, bukan jumlah slot maksimum pada template.
    """
    from openpyxl import load_workbook

    equipment = []
    wb = None
    try:
        wb = load_workbook(
            excel_copy_path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        ws = _dokpil_input_sheet(wb)
        if ws is None:
            return equipment
        for row in ws.iter_rows(min_row=4, max_row=9, min_col=2, max_col=5, values_only=True):
            _no, _jenis, _kapasitas, _jumlah = row
            if str(_jenis or "").strip().casefold() in {"jenis alat", "nama alat", "nama peralatan utama"}:
                continue
            jenis = _meaningful_dokpil_value(_jenis)
            if not jenis:
                continue
            equipment.append({
                "no": str(len(equipment) + 1),
                "jenis": jenis,
                "kapasitas": _meaningful_dokpil_value(_kapasitas),
                "jumlah": _meaningful_dokpil_value(_jumlah),
            })
    except Exception as exc:
        print(f"Warning read Dokpil equipment: {exc}")
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
    return equipment


def _read_dokpil_personnel(excel_copy_path):
    """Baca personil aktif dari sheet ``Personil!B4:F9`` atau ``G4:K9``.

    Sheet terpisah ``Personil`` memakai kolom No/Jabatan/Sertifikat/
    Pengalaman/Jumlah. Layout gabungan lama memakai kolom yang sama mulai
    dari G dan tetap kompatibel jika kolom Jumlah belum tersedia.
    Baris tanpa Jabatan (termasuk hasil formula ``0``) tidak dikirim ke Word;
    nomor dinormalisasi ulang agar tidak ada nomor kosong atau loncat.
    """
    from openpyxl import load_workbook

    personnel = []
    wb = None
    try:
        wb = load_workbook(
            excel_copy_path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        ws, standalone = _dokpil_personnel_sheet(wb)
        if ws is None:
            return personnel
        if standalone:
            rows = ws.iter_rows(min_row=4, max_row=9, min_col=2, max_col=6, values_only=True)
        else:
            rows = ws.iter_rows(min_row=4, max_row=9, min_col=7, max_col=11, values_only=True)
        for row in rows:
            _no, _jabatan, _sertifikat, _pengalaman, _jumlah = row
            if str(_jabatan or "").strip().casefold() in {"jabatan", "jabatan personil", "nama personil"}:
                continue
            jabatan = _meaningful_dokpil_value(_jabatan)
            if not jabatan:
                continue
            personnel.append({
                "no": str(len(personnel) + 1),
                "jabatan": jabatan,
                "sertifikat": _meaningful_dokpil_value(_sertifikat),
                "pengalaman": _meaningful_dokpil_value(_pengalaman),
                "jumlah": _meaningful_dokpil_value(_jumlah),
            })
    except Exception as exc:
        print(f"Warning read Dokpil personnel: {exc}")
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
    return personnel


def _set_data_aliases(data, field_name, value):
    """Set nama field asli + nama hasil normalisasi Word mail merge."""
    data[field_name] = value
    data[normalize_field_name(field_name)] = value


def _angka_kata(n):
    angka = {
        0: "Nol", 1: "Satu", 2: "Dua", 3: "Tiga", 4: "Empat",
        5: "Lima", 6: "Enam", 7: "Tujuh", 8: "Delapan", 9: "Sembilan",
    }
    return angka.get(int(n), str(int(n)))


def _augment_ba_counts(data, excel_copy_path):
    """Ambil hitungan BA dari Sheet 0 Input BA, bukan cache satu_data."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_copy_path, read_only=True, data_only=True, keep_links=False)
        if "0. Input BA" not in wb.sheetnames:
            wb.close()
            return
        ws = wb["0. Input BA"]
        total = int(ws["C25"].value or 0)
        openable = int(ws["C26"].value or 0)
        incomplete = int(ws["C28"].value or 0)
        unreadable = int(ws["C29"].value or 0)
        complete = max(total - incomplete, 0)
        values = {
            "Ket 1": f"Terdapat jumlah dokumen penawaran keseluruhan sebanyak {total} ({_angka_kata(total)}) buah;",
            "Ket 2 a": f"Jumlah dokumen penawaran yang lengkap sebanyak {complete} ({_angka_kata(complete)}) buah;",
            "Ket 2 b": f"Jumlah dokumen penawaran yang tidak lengkap sebanyak {incomplete} ({_angka_kata(incomplete)}) buah;",
            "Ket 3 a": f"Jumlah dokumen penawaran yang dapat dibuka sebanyak {openable} ({_angka_kata(openable)}) buah;",
            "Ket 3 b": f"Jumlah dokumen penawaran yang tidak dapat dibuka sebanyak {unreadable} ({_angka_kata(unreadable)}) buah;",
        }
        for key, value in values.items():
            _set_data_aliases(data, key, value)
        wb.close()
    except Exception:
        pass


def _trim_blank_participant_rows(wdDoc):
    """Hapus baris peserta kosong dari tabel ringkasan BA Pembuktian."""
    try:
        for i in range(1, wdDoc.Tables.Count + 1):
            table = wdDoc.Tables(i)
            table_text = table.Range.Text.upper()
            if "NAMA PESERTA" not in table_text or "KETERANGAN" not in table_text:
                continue
            for r in range(table.Rows.Count, 1, -1):
                row = table.Rows(r)
                first = row.Cells(1).Range.Text.replace("\r", "").replace("\a", "").strip()
                second = row.Cells(2).Range.Text.replace("\r", "").replace("\a", "").strip()
                if re.fullmatch(r"[23]\.?", first) and (not second or second == "0"):
                    row.Delete()
    except Exception:
        pass


def _blank_empty_participant_rows(wdDoc, data=None):
    """Kosongkan seluruh isi baris peserta 2/3 bila nama pesertanya kosong.

    Template ringkasan evaluasi selalu menyediakan tiga slot peserta. Formula
    Excel mengisi slot yang tidak dipakai dengan 0/MS, sehingga mengosongkan
    field nama saja masih meninggalkan angka dan status palsu di kolom lain.
    Baris dan border dipertahankan; hanya teks tiap sel yang dibersihkan.
    """
    def _cell_text(cell):
        return cell.Range.Text.replace("\r", "").replace("\a", "").strip()

    try:
        for i in range(1, wdDoc.Tables.Count + 1):
            table = wdDoc.Tables(i)
            table_text = table.Range.Text.upper()
            if "PESERTA" not in table_text or "SYARAT KUALIFIKASI" in table_text:
                continue
            if table.Rows.Count < 2:
                continue
            for r in range(2, table.Rows.Count + 1):
                row = table.Rows(r)
                first = _cell_text(row.Cells(1))
                if not re.fullmatch(r"[23]\.?", first):
                    continue
                if row.Cells.Count < 2:
                    continue
                slot = first.rstrip(".")
                expected = data.get(f"Peserta_{slot}") if data else None
                participant = _cell_text(row.Cells(2))
                data_empty = expected is not None and str(expected).strip() in (
                    "", "0", "0.0", "None", "null"
                )
                cell_empty = participant in ("", "0", "0.0", "None", "null")
                if not data_empty and not cell_empty:
                    continue
                for cell in row.Cells:
                    rng = cell.Range.Duplicate
                    rng.End = max(rng.Start, rng.End - 1)
                    rng.Text = ""
    except Exception:
        pass


def _blank_empty_participant_rows_xml(docx_path, data):
    """Bersihkan row peserta kosong langsung pada XML salinan DOCX.

    Sebagian tabel ringkasan berada pada struktur Word yang tidak selalu masuk
    ke koleksi COM ``Document.Tables``. XML menjadi lapisan deterministik
    sebelum DOCX dibuka Word; format tabel tetap, isi row saja yang kosong.
    """
    if not data:
        return
    import tempfile
    import zipfile
    tmp_path = docx_path + ".xmltmp"
    changed = False
    try:
        with zipfile.ZipFile(docx_path, "r") as zin:
            document_xml = zin.read("word/document.xml").decode("utf-8")
            def text_from_xml(fragment):
                return "".join(re.findall(r"<w:t\b[^>]*>(.*?)</w:t>", fragment, re.S))

            def clear_cell(match):
                cell = match.group(0)
                props = re.search(r"<w:tcPr\b[^>]*>.*?</w:tcPr\s*>", cell, re.S)
                prefix = props.group(0) if props else ""
                return re.sub(
                    r"(<w:tc\b[^>]*>).*?(</w:tc\s*>)",
                    lambda m: m.group(1) + prefix + "<w:p/>" + m.group(2),
                    cell, count=1, flags=re.S,
                )

            def process_table(table_match):
                nonlocal changed
                table = table_match.group(0)
                upper = text_from_xml(table).upper()
                if "PESERTA" not in upper or "SYARAT KUALIFIKASI" in upper:
                    return table
                rows = re.compile(r"<w:tr\b[^>]*>.*?</w:tr\s*>", re.S)

                def process_row(row_match):
                    nonlocal changed
                    row = row_match.group(0)
                    cells = re.findall(r"<w:tc\b[^>]*>.*?</w:tc\s*>", row, re.S)
                    if len(cells) < 2:
                        return row
                    first = text_from_xml(cells[0]).strip()
                    if not re.fullmatch(r"[23]\.?", first):
                        return row
                    slot = first.rstrip(".")
                    expected = str(data.get(f"Peserta_{slot}") or "").strip()
                    if expected not in ("", "0", "0.0", "None", "null"):
                        return row
                    new_row = re.sub(
                        r"<w:tc\b[^>]*>.*?</w:tc\s*>",
                        clear_cell, row, flags=re.S,
                    )
                    changed |= new_row != row
                    return new_row

                return rows.sub(process_row, table)

            tables = re.compile(r"<w:tbl\b[^>]*>.*?</w:tbl\s*>", re.S)
            new_xml = tables.sub(process_table, document_xml)
            if not changed:
                return
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    payload = (
                        new_xml.encode("utf-8")
                        if item.filename == "word/document.xml"
                        else zin.read(item.filename)
                    )
                    zout.writestr(item, payload)
        os.replace(tmp_path, docx_path)
    except Exception:
        try:
            os.remove(tmp_path)
        except Exception:
            pass


def _dokpil_personnel_value(data, field_name, slot):
    """Ambil field Personil N dari data list_dokpil secara toleran."""
    base = f"{field_name}_{slot}"
    for key in (base, normalize_field_name(base)):
        if key in data:
            value = data[key]
            if value is None:
                return ""
            value = format_value(value)
            return "" if value in ("", "0", "0.0", "None", "null") else value
    return ""


def _prepare_dokpil_personnel_table(wdDoc, data):
    """Sesuaikan jumlah row tenaga ahli Dokpil dengan slot Personil aktif.

    Template Dokpil lama hanya menyimpan row Personil 1. Fungsi ini mencari
    tabel persyaratan teknis, meng-clone row tersebut untuk slot 2/3 yang
    terisi, mengganti kode MERGEFIELD pada row clone, dan menghapus row
    personel ekstra bila paket hanya memiliki satu/dua tenaga ahli.
    """
    try:
        active_slots = []
        for slot in range(1, 4):
            fields = (
                "Jabatan_Personil", "Pengalaman_Kerja_Personil",
                "Sertifikat_Personil", "Minimal_Orang",
            )
            if any(_dokpil_personnel_value(data, name, slot) for name in fields):
                active_slots.append(slot)
        if not active_slots:
            return
        print(f"Dokpil active personnel slots: {active_slots}")

        target = None
        source_row = None
        field_rows = {}
        for ti in range(1, wdDoc.Tables.Count + 1):
            table = wdDoc.Tables(ti)
            text = table.Range.Text.upper()
            if "JABATAN DALAM PEKERJAAN" not in text or "SERTIFIKAT KOMPETENSI" not in text:
                continue
            # Jangan enumerasi table.Rows: header Dokpil memakai vertically
            # merged cells dan Word COM melempar error pada baris tertentu.
            for fi in range(1, table.Range.Fields.Count + 1):
                field = table.Range.Fields(fi)
                code = field.Code.Text.upper()
                match = re.search(r"PERSONIL_(\d+)", code)
                if not match:
                    continue
                try:
                    row_index = field.Result.Cells(1).RowIndex
                except Exception:
                    continue
                field_rows.setdefault(int(match.group(1)), set()).add(row_index)
            if 1 in field_rows:
                target = table
                source_row = table.Rows(next(iter(field_rows[1])))
            if target is not None:
                break
        if target is None or source_row is None:
            print("Warning dynamic Dokpil personnel row: target table/Personil 1 field tidak ditemukan")
            return

        # Hapus row tenaga ahli yang sudah ada tetapi tidak dipakai.
        existing_slots = set(field_rows)
        for slot in sorted(set(field_rows) - set(active_slots), reverse=True):
            for row_index in sorted(field_rows[slot], reverse=True):
                try:
                    target.Rows(row_index).Delete()
                except Exception:
                    pass
            existing_slots.discard(slot)

        # Row sumber dapat bergeser setelah penghapusan; ambil ulang dari
        # field Personil 1 yang masih ada.
        for fi in range(1, target.Range.Fields.Count + 1):
            field = target.Range.Fields(fi)
            if "PERSONIL_1" in field.Code.Text.upper():
                source_row = target.Rows(field.Result.Cells(1).RowIndex)
                break
        if source_row is None:
            print("Warning dynamic Dokpil personnel row: source row hilang setelah cleanup")
            return

        # Clone row Personil 1 untuk setiap slot yang belum ada.
        for slot in active_slots:
            if slot == 1 or slot in existing_slots:
                continue
            source_row.Range.Copy()
            new_row = target.Rows.Add()
            new_row.Range.Paste()
            for fi in range(new_row.Range.Fields.Count, 0, -1):
                field = new_row.Range.Fields(fi)
                code = field.Code.Text
                if "PERSONIL_1" in code:
                    field.Code.Text = code.replace("PERSONIL_1", f"PERSONIL_{slot}")
            cell_rng = new_row.Cells(1).Range
            cell_rng.End = cell_rng.End - 1  # pertahankan end-of-cell marker
            cell_rng.Text = f"{slot}."
            existing_slots.add(slot)
        print(f"Dokpil personnel table ready: {len(active_slots)} row(s)")
    except Exception as exc:
        # Dokpil tetap boleh dicetak dengan template asli bila Word menolak
        # operasi row/clipboard pada versi Office tertentu.
        import traceback
        print(f"Warning dynamic Dokpil personnel row: {exc}")
        traceback.print_exc()


def _prepare_dokpil_personnel_markers_docx(docx_path, personnel):
    """Isi tabel personil berbasis marker ``[[..._PERSONIL]]``.

    Marker dipasang pada satu row donor di tabel Word. Row donor dihapus,
    lalu dikloning sesuai personil aktif dari Excel. Iterasi hanya memproses
    row langsung pada tabel target agar marker di nested table tidak salah
    dianggap sebagai row tabel pembungkus.
    """
    import copy as _copy
    import zipfile
    from lxml import etree as LET

    ns = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    marker_fields = {
        "[[NO_PERSONIL]]": "no",
        "[[JABATAN_PERSONIL]]": "jabatan",
        "[[JUMLAH_PERSONIL]]": "jumlah",
        "[[SERTIFIKAT_PERSONIL]]": "sertifikat",
        "[[PENGALAMAN_PERSONIL]]": "pengalaman",
    }
    changed = False

    with zipfile.ZipFile(docx_path, "r") as zin:
        names = zin.namelist()
        if "word/document.xml" not in names:
            return False
        root = LET.fromstring(zin.read("word/document.xml"))

        targets = []
        for table in root.iter(ns + "tbl"):
            rows = [child for child in table if child.tag == ns + "tr"]
            for index, row in enumerate(rows):
                cells = [child for child in row if child.tag == ns + "tc"]
                if not cells:
                    continue
                # Row pembungkus dapat memuat nested table marker; target
                # sebenarnya akan ditemukan pada iterasi tabel anak.
                if row.find(".//" + ns + "tbl") is not None:
                    continue
                row_text = _xml_text_content(row, ns)
                if not any(marker in row_text for marker in marker_fields):
                    continue
                fields_by_column = []
                for cell in cells:
                    cell_text = _xml_text_content(cell, ns)
                    fields_by_column.append(
                        next(
                            (field for marker, field in marker_fields.items() if marker in cell_text),
                            None,
                        )
                    )
                if any(fields_by_column):
                    targets.append((table, index, row, fields_by_column))

        if not targets:
            return False

        for target, _index, donor, fields_by_column in targets:
            # Posisi donor disimpan sebelum row marker dihapus.
            insert_at = list(target).index(donor)
            target.remove(donor)
            for item in personnel or []:
                clone = _copy.deepcopy(donor)
                cells = [child for child in clone if child.tag == ns + "tc"]
                for cell, field in zip(cells, fields_by_column):
                    if field:
                        _set_dokpil_cell_text(
                            cell,
                            item.get(field, ""),
                            ns,
                            LET,
                            _copy,
                        )
                target.insert(insert_at, clone)
                insert_at += 1
            changed = True

        if not changed:
            return False
        new_xml = LET.tostring(root, encoding="UTF-8", xml_declaration=True, standalone=True)
        tmp = docx_path + ".tmp"
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in names:
                zout.writestr(
                    name,
                    new_xml if name == "word/document.xml" else zin.read(name),
                )
    os.replace(tmp, docx_path)
    print(f"Dokpil personnel marker table ready: {len(targets)} table(s), {len(personnel or [])} row(s) each")
    return True


def _prepare_dokpil_personnel_docx(docx_path, data):
    """Clone row tenaga ahli pada XML DOCX sebelum Word membukanya.

    Template Dokpil memakai vertically merged cells sehingga Word COM tidak
    dapat mengakses ``Rows(n)`` secara konsisten. XML DOCX tetap deterministik:
    row yang berisi MERGEFIELD Jabatan_Personil_1 dicopy untuk slot aktif 2/3.
    Hanya file (Merged) sementara yang diubah; template asli aman.
    """
    import copy as _copy
    import zipfile
    from lxml import etree as LET

    # Jalur baru: marker kurung siku membaca tabel input G:J. Jika template
    # belum memakai marker baru, lanjutkan ke kompatibilitas MERGEFIELD lama.
    personnel = (data or {}).get("_dokpil_personnel")
    if personnel is not None and _prepare_dokpil_personnel_markers_docx(docx_path, personnel):
        return

    active_slots = []
    for slot in range(1, 4):
        if any(_dokpil_personnel_value(data, name, slot) for name in (
            "Jabatan_Personil", "Pengalaman_Kerja_Personil",
            "Sertifikat_Personil", "Minimal_Orang",
        )):
            active_slots.append(slot)
    if len(active_slots) <= 1:
        return

    ns = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    with zipfile.ZipFile(docx_path, "r") as zin:
        names = zin.namelist()
        if "word/document.xml" not in names:
            return
        root = LET.fromstring(zin.read("word/document.xml"))

        source = None
        parent = None
        # Ada outer row yang membungkus nested table header+data. Pilih row
        # terdalam (tepat satu w:tr di dalamnya), yaitu row data sebenarnya.
        candidates = []
        parent_map = {child: p for p in root.iter() for child in list(p)}
        for candidate in root.iter(ns + "tr"):
            text = "".join(t.text or "" for t in candidate.iter() if t.tag in (ns + "t", ns + "instrText"))
            if "JABATAN_PERSONIL_1" not in text.upper():
                continue
            nested_rows = list(candidate.iter(ns + "tr"))
            candidates.append((len(nested_rows), candidate))
        if candidates:
            _, source = min(candidates, key=lambda item: item[0])
            parent = parent_map.get(source)
        if source is None or parent is None:
            return

        rows = [child for child in list(parent) if child.tag == ns + "tr"]
        # Hapus clone personel yang tidak dipakai, bila template sudah pernah
        # diproses sebelumnya.
        for row in rows:
            text = "".join(t.text or "" for t in row.iter() if t.tag in (ns + "t", ns + "instrText"))
            match = re.search(r"PERSONIL_(\d+)", text.upper())
            if match and int(match.group(1)) not in active_slots:
                parent.remove(row)

        # Refresh posisi source setelah penghapusan.
        source_index = list(parent).index(source)
        existing = set()
        for row in list(parent):
            if row.tag != ns + "tr":
                continue
            text = "".join(t.text or "" for t in row.iter() if t.tag in (ns + "t", ns + "instrText"))
            for match in re.finditer(r"PERSONIL_(\d+)", text.upper()):
                existing.add(int(match.group(1)))

        insert_at = source_index + 1
        for slot in active_slots:
            if slot == 1 or slot in existing:
                continue
            clone = _copy.deepcopy(source)
            for node in clone.iter():
                if node.tag not in (ns + "t", ns + "instrText") or not node.text:
                    continue
                node.text = re.sub(
                    r"(?i)(Jabatan|Pengalaman_Kerja|Sertifikat|Minimal_Orang|Personil)_1",
                    lambda match: match.group(0)[:-1] + str(slot),
                    node.text,
                )
            cells = [c for c in clone if c.tag == ns + "tc"]
            if cells:
                for node in cells[0].iter():
                    if node.tag == ns + "t" and (node.text or "").strip() in ("1.", "1"):
                        node.text = node.text.replace("1", str(slot), 1)
                        break
            parent.insert(insert_at, clone)
            insert_at += 1
            existing.add(slot)

        new_xml = LET.tostring(root, encoding="UTF-8", xml_declaration=True, standalone=True)
        tmp = docx_path + ".tmp"
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in names:
                zout.writestr(name, new_xml if name == "word/document.xml" else zin.read(name))
    os.replace(tmp, docx_path)
    print(f"Dokpil personnel XML ready: {len(active_slots)} row(s)")


def _xml_text_content(node, ns):
    """Ambil teks visible + kode merge field dari node Word XML."""
    text_tags = {ns + "t", ns + "instrText", ns + "delText"}
    return "".join(
        child.text or ""
        for child in node.iter()
        if child.tag in text_tags
    )


def _set_dokpil_cell_text(cell, value, ns, LET, copy_module):
    """Ganti isi sel tanpa membuang tcPr/pPr/rPr donor."""
    paragraphs = [child for child in cell if child.tag == ns + "p"]
    if paragraphs:
        paragraph = paragraphs[0]
    else:
        paragraph = LET.Element(ns + "p")
        cell.append(paragraph)

    paragraph_properties = paragraph.find(ns + "pPr")
    run_properties = None
    first_run = paragraph.find(".//" + ns + "r")
    if first_run is not None:
        first_run_properties = first_run.find(ns + "rPr")
        if first_run_properties is not None:
            run_properties = copy_module.deepcopy(first_run_properties)

    # Row donor sudah membawa paragraph/cell formatting. Bersihkan hanya
    # isi paragraph agar border, alignment, font, dan width tetap utuh.
    for child in list(paragraph):
        if child is not paragraph_properties:
            paragraph.remove(child)

    run = LET.SubElement(paragraph, ns + "r")
    if run_properties is not None:
        run.append(run_properties)
    text_node = LET.SubElement(run, ns + "t")
    value = "" if value is None else str(value)
    if value[:1].isspace() or value[-1:].isspace():
        text_node.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    text_node.text = value


def _prepare_dokpil_equipment_docx(docx_path, data):
    """Isi tabel marker alat dari sheet input alat/personil.

    Target dikenali dari marker row, header empat kolom PLPK, atau header
    delapan kolom Dokpil Tender. Marker canonical mengikuti template
    Dokpil/Isi Reviu:
    ``[[NO_ALAT]]``, ``[[NAMA_ALAT]]``, ``[[JUMLAH_ALAT]]``,
    ``[[KAPASITAS_ALAT]]``. Semua perubahan hanya dilakukan pada salinan
    ``(Merged)``.
    """
    import copy as _copy
    import zipfile
    from lxml import etree as LET

    equipment = data.get("_dokpil_equipment") if data else None
    if equipment is None:
        return

    ns = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    marker_fields = {
        "[[NO_ALAT]]": "no",
        "[[NAMA_ALAT]]": "jenis",
        "[[JUMLAH_ALAT]]": "jumlah",
        "[[KAPASITAS_ALAT]]": "kapasitas",
        # Marker tambahan tabel Tender. Sumber Excel saat ini hanya
        # menyediakan empat field canonical di atas.
        "[[MERK_TIPE_ALAT]]": None,
        "[[KONDISI_ALAT]]": None,
        "[[STATUS_KEPEMILIKAN_ALAT]]": None,
        "[[KETERANGAN_ALAT]]": None,
    }
    header_fields = {
        "no": "no",
        "jenis": "jenis",
        "namaalat": "jenis",
        "namaperalatan": "jenis",
        "namaperalatanutama": "jenis",
        "jenisalat": "jenis",
        "jumlah": "jumlah",
        "jumlahunitbuah": "jumlah",
        "kapasitas": "kapasitas",
        "kapasitasminimal": "kapasitas",
    }
    with zipfile.ZipFile(docx_path, "r") as zin:
        names = zin.namelist()
        if "word/document.xml" not in names:
            return
        root = LET.fromstring(zin.read("word/document.xml"))

        targets = []
        for table in root.iter(ns + "tbl"):
            rows = [child for child in table if child.tag == ns + "tr"]
            for index, row in enumerate(rows):
                cells = [child for child in row if child.tag == ns + "tc"]
                values = [
                    re.sub(r"\s+", " ", _xml_text_content(cell, ns)).strip().casefold()
                    for cell in cells
                ]
                normalized_headers = [
                    re.sub(r"[^a-z0-9]+", "", value) for value in values
                ]
                fields_by_column = [header_fields.get(value) for value in normalized_headers]
                row_text = _xml_text_content(row, ns)
                # Dokpil membungkus tabel alat di dalam sel tabel persyaratan
                # teknis. Jangan salah menganggap row pembungkus sebagai
                # target hanya karena teks marker ikut terbaca dari nested
                # table; target sebenarnya ditemukan saat iterasi tabel anak.
                has_nested_table = row.find(".//" + ns + "tbl") is not None
                marker_row = (
                    not has_nested_table
                    and any(marker in row_text for marker in marker_fields)
                )
                canonical_header_ok = (
                    len(fields_by_column) == 4
                    and set(fields_by_column) == {"no", "jenis", "jumlah", "kapasitas"}
                )
                tender_header_ok = (
                    len(normalized_headers) >= 5
                    and (
                        normalized_headers[:5] == [
                            "no", "namaperalatanutama", "merkdantipe", "kapasitas", "jumlah"
                        ]
                        or normalized_headers[:6] == [
                            "no", "jenis", "merekdantipe", "kapasitas", "jumlah", "kepemilikanstatus"
                        ]
                    )
                )
                header_row_ok = canonical_header_ok or tender_header_ok
                if marker_row or header_row_ok:
                    if marker_row:
                        marker_fields_by_column = []
                        for cell in cells:
                            cell_text = _xml_text_content(cell, ns)
                            marker = next(
                                (marker for marker in marker_fields if marker in cell_text),
                                None,
                            )
                            marker_fields_by_column.append(
                                marker_fields[marker] if marker is not None else None
                            )
                        fields_by_column = marker_fields_by_column
                    if index + 1 < len(rows):
                        targets.append((table, index, row, rows[index + 1], fields_by_column))
                    break

        if not targets:
            print("Warning dynamic equipment row: target table/header tidak ditemukan")
            return

        for target, header_index, header_row, source_row, fields_by_column in targets:
            rows = [child for child in target if child.tag == ns + "tr"]
            # Template produksi memiliki satu row donor. Hapus seluruh row
            # data agar hasil tidak menggandakan alat saat template sudah
            # pernah diisi. Hanya tabel yang memiliki marker/header yang masuk.
            for row in rows[header_index + 1:]:
                target.remove(row)

            insert_at = list(target).index(header_row) + 1
            # source_row sudah dilepas dari target; salin formatnya untuk
            # setiap alat aktif. Tanpa alat, tabel dibiarkan header-only.
            for item in equipment:
                clone = _copy.deepcopy(source_row)
                cells = [child for child in clone if child.tag == ns + "tc"]
                if len(cells) < len(fields_by_column):
                    print("Warning dynamic equipment row: jumlah kolom donor tidak cocok")
                    break
                item_values = {
                    "no": item.get("no", ""),
                    "jenis": item.get("jenis", ""),
                    "jumlah": item.get("jumlah", ""),
                    "kapasitas": item.get("kapasitas", ""),
                }
                for cell, field in zip(cells, fields_by_column):
                    value = item_values.get(field, "")
                    _set_dokpil_cell_text(cell, value, ns, LET, _copy)
                target.insert(insert_at, clone)
                insert_at += 1

        new_xml = LET.tostring(root, encoding="UTF-8", xml_declaration=True, standalone=True)
        tmp = docx_path + ".tmp"
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in names:
                zout.writestr(
                    name,
                    new_xml if name == "word/document.xml" else zin.read(name),
                )
    os.replace(tmp, docx_path)
    print(f"Equipment table(s) ready: {len(targets)} table(s), {len(equipment)} row(s) each")


def cleanup_blank_pages(doc):
    # Metode Ringan & Cepat: Shrink paragraf kosong di seluruh dokumen (tanpa batas halaman).
    try:
        for i, para in enumerate(doc.Paragraphs):
            try:
                txt = para.Range.Text.replace('\r', '').replace('\n', '')
                # Hanya panggil .Information(3) (yang butuh komputasi layout berat) JIKA paragraf kosong.
                # Ini mempercepat script 10x lipat dan mencegah hang/timeout dari RPC Server!
                if not txt.strip():
                    pg_num = para.Range.Information(3) # wdActiveEndPageNumber
                    if pg_num > 1:
                        # Shrink blank space (Enter Berlebihan di Word)
                        para.Range.Font.Size = 1
                        para.Format.SpaceBefore = 0
                        para.Format.SpaceAfter = 0
                        para.Format.LineSpacingRule = 0
            except:
                pass
    except Exception as e:
        print(f"Warning saat cleanup: {e}")
        pass
    
    return


def _fit_path(folder, filename, max_total=240):
    """Word COM (ExportAsFixedFormat/SaveAs2) menolak path >255 char
    ('String is longer than 255 characters', wdmain11.chm 41873).
    Potong stem filename agar total path aman."""
    path = os.path.join(folder, filename)
    if len(path) <= max_total:
        return path
    stem, ext = os.path.splitext(filename)
    avail = max_total - len(folder) - len(ext) - 1  # -1 utk separator
    if avail < 10:
        avail = 10
    return os.path.join(folder, stem[:avail].rstrip() + ext)


def _next_available_pdf_path(target_path):
    """Pilih nama PDF kosong: target asli, lalu ``_v2``, ``_v3``, dst."""
    target_path = os.fspath(target_path)
    if not os.path.exists(target_path):
        return target_path

    stem, ext = os.path.splitext(target_path)
    version = 2
    while True:
        candidate = f"{stem}_v{version}{ext}"
        if not os.path.exists(candidate):
            return candidate
        version += 1


def _strip_pl_ba_signature_header(wd_doc) -> bool:
    """Hapus header ter-link pada section lembar tanda tangan BA Reviu PL.

    Template BA Reviu PLJKK memakai section baru untuk lembar 3, tetapi header
    section tersebut masih ``LinkToPrevious``. Akibatnya header instansi dari
    halaman 1-2 ikut tercetak di halaman tanda tangan. Perubahan hanya berlaku
    pada salinan Word sementara yang akan diekspor/print.
    """
    try:
        if int(wd_doc.ComputeStatistics(2)) < 3 or wd_doc.Sections.Count < 2:
            return False
        page_two = wd_doc.GoTo(1, 1, 2)  # wdGoToPage, wdGoToAbsolute
        page_three = wd_doc.GoTo(1, 1, 3)
        section_two = int(page_two.Information(2))  # wdActiveEndSectionNumber
        section_three = int(page_three.Information(2))
        if section_three <= section_two:
            return False

        header_section = wd_doc.Sections(section_three)
        for header_kind in (1, 2, 3):  # primary, first-page, even-page
            header = header_section.Headers(header_kind)
            header.LinkToPrevious = False
            header.Range.Text = ""
        return True
    except Exception as exc:
        print(f"Warning: header lembar tanda tangan PL tidak dihapus: {exc}")
        return False


def _word_process_id(word_app, word_doc=None):
    """Ambil PID instance Word yang dibuat DispatchEx; None jika gagal."""
    import ctypes
    for owner in (word_app, getattr(word_app, "ActiveWindow", None), word_doc,
                  getattr(word_doc, "ActiveWindow", None) if word_doc is not None else None):
        try:
            hwnd = int(getattr(owner, "Hwnd"))
            pid = ctypes.c_ulong(0)
            ctypes.windll.user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
            if pid.value:
                return int(pid.value)
        except Exception:
            continue
    return None


def _terminate_word_process(pid):
    """Tutup paksa hanya PID Word milik engine, bukan Word user lain."""
    if not pid:
        return
    try:
        import subprocess
        subprocess.run(
            ["taskkill", "/PID", str(pid), "/T", "/F"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
    except Exception:
        pass


def _strip_mailmerge_datasource(docx_path):
    """Hapus attachment mail merge (w:mailMerge di word/settings.xml) dari copy
    (Merged). Path Excel panjang bikin connection string/SQL >255 char sehingga
    Word error 41873 'String is longer than 255 characters' saat auto-connect
    data source di Documents.Open. Script ini merge field sendiri via COM,
    jadi attachment tidak diperlukan di file copy."""
    import re
    import zipfile
    try:
        with zipfile.ZipFile(docx_path, "r") as zin:
            names = zin.namelist()
            if "word/settings.xml" not in names:
                return
            settings = zin.read("word/settings.xml").decode("utf-8")
            new_settings = re.sub(
                r"<w:mailMerge>.*?</w:mailMerge>|<w:mailMerge\s*/>",
                "", settings, flags=re.DOTALL)
            if new_settings == settings:
                return
            items = [(n, zin.read(n)) for n in names]
        tmp = docx_path + ".tmp"
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for n, blob in items:
                zout.writestr(n, new_settings.encode("utf-8") if n == "word/settings.xml" else blob)
        os.replace(tmp, docx_path)
    except Exception:
        pass  # gagal strip -> lanjut; worst case error lama muncul lagi


def _set_field_result(field, val):
    if len(val) <= 255:
        field.Result.Text = val
        field.Unlink()
        return
    rng = field.Result
    field.Unlink()
    rng.Text = val


def _replace_merge_fields(wdDoc, data):
    """Replace semua MERGEFIELD di wdDoc dgn nilai dari data + apply format switch.
    Field di-Unlink jadi teks statis. Loop backwards supaya index aman."""
    field_count = wdDoc.Fields.Count
    for i in range(field_count, 0, -1):
        try:
            field = wdDoc.Fields(i)
            code_text = field.Code.Text.strip()
            if code_text.upper().startswith("MERGEFIELD"):
                parts = code_text.split()
                if len(parts) >= 2:
                    fname = parts[1].strip('"').strip()
                    val = None
                    if fname in data:
                        val = data[fname]
                    else:
                        norm = normalize_field_name(fname)
                        if norm in data:
                            val = data[norm]

                    if val is not None:
                        val = str(val)
                        format_str = " ".join(parts[2:]).upper()
                        if "UPPER" in format_str:
                            val = val.upper()
                        elif "LOWER" in format_str:
                            val = val.lower()
                        elif "FIRSTCAP" in format_str:
                            val = val.capitalize()
                    else:
                        val = ""
                    _set_field_result(field, val)
        except:
            pass


def _toc_line_key(text):
    """Normalisasi label TOC tanpa mengubah teks/format aslinya."""
    import re
    return re.sub(r"\s+", " ", str(text or "").replace("\r", "")).strip().casefold()


def _toc_field_label(field):
    """Ambil label sebelum tab nomor halaman dari field PAGEREF."""
    try:
        line = str(field.Result.Paragraphs(1).Range.Text).rstrip("\r")
        if "\t" in line:
            line = line.rsplit("\t", 1)[0]
        return _toc_line_key(line)
    except Exception:
        return ""


def _update_toc_fields(wdDoc):
    """Refresh page results without rebuilding the template TOC.

    ``TablesOfContents.Update`` pada template PLPK menghapus entry manual
    BAB I dan menggeser right tab stop. Word sendiri sudah memperbarui
    PAGEREF yang memiliki target saat repagination/export. Satu-satunya
    target stale yang perlu ditangani adalah field PAGEREF tanpa bookmark;
    field tersebut di-unlink menjadi teks cached agar tidak menghasilkan
    ``ERROR! BOOKMARK NOT DEFINED.``.
    """
    try:
        toc_count = wdDoc.TablesOfContents.Count
    except Exception:
        toc_count = 0
    if not toc_count:
        try:
            wdDoc.Fields.Update()
        except Exception:
            pass
        return

    try:
        wdDoc.Bookmarks.ShowHidden = True
    except Exception:
        pass
    try:
        toc = wdDoc.TablesOfContents(1)
        fields = toc.Range.Fields
        unlinked = 0
        import re
        for index in range(fields.Count, 0, -1):
            field = fields(index)
            try:
                if int(field.Type) != 37:  # wdFieldPageRef
                    continue
                tokens = str(field.Code.Text).split()
                bookmark = tokens[1] if len(tokens) > 1 else ""
                if not bookmark or wdDoc.Bookmarks.Exists(bookmark):
                    continue
                old_result = str(field.Result.Text)
                if not re.search(r"\d+", old_result):
                    label = _toc_field_label(field)
                    if label.startswith("bab i."):
                        old_result = "- 5 -"
                field.Result.Text = old_result
                field.Unlink()
                unlinked += 1
            except Exception:
                continue
        print(f"TOC layout preserved; stale page field(s) unlinked: {unlinked}")
    except Exception as exc:
        print(f"Warning refresh TOC page results: {exc}")


def _protect_signature_layout(wdDoc):
    """Jaga blok tanda tangan tetap utuh saat Word melakukan pagination."""
    for i in range(1, wdDoc.Tables.Count + 1):
        try:
            table = wdDoc.Tables(i)
            text = table.Range.Text.upper()
            if not any(marker in text for marker in (
                "DIREKTUR/PIMPINAN", "KELOMPOK KERJA PEMILIHAN"
            )):
                continue

            # Nama tenaga ahli/pimpinan bisa membuat blok ini tinggi. Jangan
            # izinkan Word memecah baris atau mendorong label dan nama ke page
            # berbeda; Word akan memindahkan blok utuh ke halaman berikutnya.
            for j in range(1, table.Rows.Count + 1):
                row = table.Rows(j)
                try:
                    row.AllowBreakAcrossPages = False
                    row.HeightRule = 0  # wdRowHeightAuto
                except Exception:
                    pass

            paragraphs = table.Range.Paragraphs
            for j in range(1, paragraphs.Count + 1):
                paragraph = paragraphs(j)
                paragraph.Range.ParagraphFormat.KeepTogether = True
                paragraph.Range.ParagraphFormat.KeepWithNext = j < paragraphs.Count

            # Blok tanda tangan yang ditempatkan di sisa ruang halaman sering
            # membuat baris terakhir (biasanya anggota Pokja terakhir atau
            # direktur) terdorong ke halaman berikutnya.  Jangan hanya
            # mengunci baris; mulai blok pada halaman baru agar seluruh blok
            # tetap utuh dalam satu halaman.
            # Jangan sisipkan page break manual. Word sudah mampu memindahkan
            # blok ke halaman berikutnya; page break tambahan dapat membuat
            # halaman header-only. Proteksi row/paragraph di atas cukup untuk
            # mencegah tanda tangan terbelah.
        except Exception:
            pass


def _sisip_2ba_pljkk(pdf_path, folder):
    """
    Sisip 2 file BA (Evaluasi /05/ + Hasil /07/) ke dalam BA_PLJKK final.

    Posisi (anchor by judul section, occurrence-aware):
      - BA Evaluasi  → setelah halaman judul 'DAFTAR HADIR PEMBUKTIAN KUALIFIKASI'
                       occurrence #1 (di antara 2 daftar hadir pembuktian).
      - BA Hasil     → setelah halaman judul 'DAFTAR HADIR KLARIFIKASI DAN NEGOSIASI'
                       occurrence #1 (di antara 2 daftar hadir klarifikasi).

    File BA dicari via prefix '1. BA Evaluasi*.pdf' & '2. BA Hasil*.pdf' di folder.
    Best-effort: jika file BA tidak ada / anchor tidak ketemu, lewati tanpa error.
    Idempotent: jika anchor occurrence tidak cukup, BA terkait dilewati.
    """
    try:
        import glob as _glob
        import pdfplumber
        from pypdf import PdfReader, PdfWriter

        _ev = sorted(_glob.glob(os.path.join(folder, "1. BA Evaluasi*.pdf")))
        _hs = sorted(_glob.glob(os.path.join(folder, "2. BA Hasil*.pdf")))
        _ev_path = _ev[0] if _ev else None
        _hs_path = _hs[0] if _hs else None
        if not _ev_path and not _hs_path:
            return  # tidak ada file BA, lewati

        _rdr = PdfReader(pdf_path)
        _n = len(_rdr.pages)

        # Identifikasi halaman ANCHOR (judul section, bukan kop/teks berulang).
        # Valid jika: judul muncul di AWAL halaman (idx < 200, setelah kop dinas) DAN
        # halaman bukan "BERITA ACARA ..." (halaman BA punya judul section di bawah).
        _PEMB = "DAFTAR HADIR PEMBUKTIAN KUALIFIKASI"
        _KLAR = "DAFTAR HADIR KLARIFIKASI DAN NEGOSIASI"
        _pemb_pages = []    # index halaman daftar hadir pembuktian
        _klarif_pages = []  # index halaman daftar hadir klarifikasi & negosiasi
        with pdfplumber.open(pdf_path) as _plb:
            for _i, _pp in enumerate(_plb.pages):
                _u = (_pp.extract_text() or "").upper()
                _is_ba = "BERITA ACARA" in _u
                _ip = _u.find(_PEMB)
                _ik = _u.find(_KLAR)
                if _ip != -1 and _ip < 200 and not _is_ba:
                    _pemb_pages.append(_i)
                if _ik != -1 and _ik < 200 and not _is_ba:
                    _klarif_pages.append(_i)

        # Titik sisip (0-based index halaman SETELAH mana BA disisipkan).
        # Evaluasi: setelah occurrence #1 pembuktian → butuh >=2 occurrence.
        _insert_after = {}  # {page_index: [list_pdf_path]}
        if _ev_path and len(_pemb_pages) >= 2:
            _insert_after.setdefault(_pemb_pages[0], []).append(_ev_path)
        if _hs_path and len(_klarif_pages) >= 2:
            _insert_after.setdefault(_klarif_pages[0], []).append(_hs_path)

        if not _insert_after:
            return  # tidak ada anchor valid, biarkan PDF apa adanya

        _writer = PdfWriter()
        for _i in range(_n):
            _writer.add_page(_rdr.pages[_i])
            for _bp in _insert_after.get(_i, []):
                try:
                    _brdr = PdfReader(_bp)
                    for _bpg in _brdr.pages:
                        _writer.add_page(_bpg)
                except Exception:
                    pass

        # Tulis ke file sementara lalu ganti (hindari korup jika gagal di tengah)
        _tmp = pdf_path + "_withba.pdf"
        with open(_tmp, "wb") as _f:
            _writer.write(_f)
        os.replace(_tmp, pdf_path)
    except Exception:
        pass  # best-effort, jangan gagalkan cetak utama


def _export_sheet_pdf(excel_path, sheet_name, out_pdf, landscape=True, fit_wide=None, fit_tall=None):
    """Export 1 sheet Excel -> PDF. Return True jika sukses, False jika sheet tak ada/gagal."""
    import win32com.client
    xlApp = None
    wb_xl = None
    try:
        xlApp = win32com.client.DispatchEx("Excel.Application")
        xlApp.Visible = False
        xlApp.DisplayAlerts = False
        wb_xl = xlApp.Workbooks.Open(excel_path, ReadOnly=True)
        try:
            ws = wb_xl.Sheets(sheet_name)
        except Exception:
            return False
        if landscape:
            ws.PageSetup.Orientation = 2  # xlLandscape
        if fit_wide is not None or fit_tall is not None:
            ws.PageSetup.Zoom = False
            if fit_wide is not None:
                ws.PageSetup.FitToPagesWide = fit_wide
            if fit_tall is not None:
                ws.PageSetup.FitToPagesTall = fit_tall
        ws.ExportAsFixedFormat(0, out_pdf)  # 0 = xlTypePDF
        return True
    except Exception:
        return False
    finally:
        if wb_xl:
            try: wb_xl.Close(False)
            except Exception: pass
        if xlApp:
            try: xlApp.Quit()
            except Exception: pass


def _stitch_excel_at_anchor(word_pdf, anchor_excel_pairs, out_pdf):
    """
    Sisip PDF Excel ke word_pdf SETELAH tiap halaman yang mengandung anchor teks.

    anchor_excel_pairs: list of (anchor_text_upper, excel_pdf_path).
      Tiap occurrence anchor di word_pdf -> sisip excel_pdf_path setelah halaman itu.
      Anchor dicocokkan berurutan: occurrence ke-N halaman anchor -> excel ke-N (jika anchor
      sama, daftarkan pasangan itu sekali per occurrence — lihat caller).

    Implementasi: scan tiap halaman word, untuk tiap anchor yang match di halaman,
    jadwalkan sisip excel-nya setelah halaman tsb. Robust thd geseran halaman.
    """
    import pdfplumber
    from pypdf import PdfReader, PdfWriter

    rdr_word = PdfReader(word_pdf)
    n_word = len(rdr_word.pages)

    # cache reader excel per path
    _excel_readers = {}
    def _rdr_excel(p):
        if p not in _excel_readers:
            _excel_readers[p] = PdfReader(p)
        return _excel_readers[p]

    # Hitung occurrence per anchor; tiap (anchor,excel) dipakai sekali berurutan.
    # Bangun antrian per anchor_text -> list excel_pdf (FIFO).
    from collections import defaultdict, deque
    queues = defaultdict(deque)
    for atext, epath in anchor_excel_pairs:
        queues[atext.upper()].append(epath)

    # Tentukan halaman -> list excel yang disisip setelahnya.
    insert_after = defaultdict(list)  # page_idx -> [excel_path,...]
    with pdfplumber.open(word_pdf) as plb:
        for pi, pp in enumerate(plb.pages):
            up = (pp.extract_text() or "").upper()
            for atext, q in queues.items():
                if q and atext in up:
                    insert_after[pi].append(q.popleft())

    writer = PdfWriter()
    for pi in range(n_word):
        writer.add_page(rdr_word.pages[pi])
        for epath in insert_after.get(pi, []):
            for epg in _rdr_excel(epath).pages:
                writer.add_page(epg)
    return _safe_write_pdf(writer, out_pdf)


def _build_bapljkk_final_pdf(wd_doc, folder, kode):
    """Export BA Word + 2 copy sheet 7.2 lalu sisipkan Summary SPSE."""
    from pypdf import PdfReader, PdfWriter
    import pdfplumber
    import win32com.client

    pdf_path = _fit_path(folder, f"BA_PLJKK_{kode}.pdf")
    tmp_word = pdf_path + "_tmpword.pdf"
    tmp_72 = pdf_path + "_tmp72.pdf"
    xlsm_paths = glob.glob(os.path.join(folder, "*.xlsm"))
    xlsm_path = os.path.normpath(xlsm_paths[0]) if xlsm_paths else None
    has_72 = False

    try:
        start = wd_doc.Sections(3).Range.Start if wd_doc.Sections.Count >= 3 else wd_doc.Content.Start
        wd_doc.Range(start, wd_doc.Content.End).ExportAsFixedFormat(
            OutputFileName=tmp_word, ExportFormat=17
        )

        if xlsm_path:
            xl_app = None
            wb = None
            try:
                xl_app = win32com.client.DispatchEx("Excel.Application")
                xl_app.Visible = False
                wb = xl_app.Workbooks.Open(xlsm_path, ReadOnly=True)
                wb.Sheets("7.2 Dengan Nego").ExportAsFixedFormat(
                    Type=0, Filename=tmp_72, Quality=0,
                    IncludeDocProperties=True, IgnorePrintAreas=False,
                    OpenAfterPublish=False,
                )
                has_72 = True
            except Exception:
                pass
            finally:
                if wb:
                    try: wb.Close(False)
                    except Exception: pass
                if xl_app:
                    try: xl_app.Quit()
                    except Exception: pass

        if has_72:
            rdr_word = PdfReader(tmp_word)
            split_page = len(rdr_word.pages)
            with pdfplumber.open(tmp_word) as pdf:
                for page_index, page in enumerate(pdf.pages):
                    if "DAFTAR HADIR KLARIFIKASI DAN NEGOSIASI" in (page.extract_text() or "").upper():
                        split_page = page_index
                        break
            rdr_72 = PdfReader(tmp_72)
            writer = PdfWriter()
            for page_index in range(split_page):
                writer.add_page(rdr_word.pages[page_index])
            for _ in range(2):
                for page in rdr_72.pages:
                    writer.add_page(page)
            for page_index in range(split_page, len(rdr_word.pages)):
                writer.add_page(rdr_word.pages[page_index])
            with open(pdf_path, "wb") as output:
                writer.write(output)
        else:
            shutil.move(tmp_word, pdf_path)
    finally:
        for tmp_path in (tmp_word, tmp_72):
            try: os.remove(tmp_path)
            except Exception: pass

    try:
        from gabung_ba_pljkk import gabung as gabung_ba_pljkk
        result = gabung_ba_pljkk(folder)
        if result.get("ok"):
            return result["output"]
    except Exception:
        pass
    return pdf_path


def merge_word(word_path, data, mode="buka", pdf_name=""):
    import pythoncom
    import win32com.client
    import glob as _glob_excel

    # `merge_word()` juga dipanggil langsung oleh beberapa workflow, bukan
    # hanya melalui CLI yang kebetulan memiliki variabel global excel_path.
    # Resolve workbook dari folder Word agar export sheet BA selalu deterministik.
    _excel_candidates = sorted(_glob_excel.glob(
        os.path.join(os.path.dirname(os.path.abspath(word_path)), "*.xlsm")
    ))
    excel_path = _excel_candidates[0] if _excel_candidates else ""

    # Mode bapljkk: copy template -> (Merged), replace MERGEFIELD dari Excel, baru export.
    # (sebelumnya buka ReadOnly tanpa merge -> PDF tampil cached hasil merge lama/template)
    if mode in ("pdf_bapljkk", "printer_bapljkk"):
        _word_path_win = os.path.abspath(os.path.normpath(word_path))
        _folder = os.path.dirname(_word_path_win)
        _base_b, _ext_b = os.path.splitext(os.path.basename(_word_path_win))
        if _ext_b.lower() not in (".docx", ".docm"):
            _ext_b = ".docx"
        _merged_b = _fit_path(_folder, f"{_base_b[:60].rstrip()} (Merged){_ext_b}")
        shutil.copy2(_word_path_win, _merged_b)
        # Header resmi dipilih dari profil instansi, hanya pada copy sementara.
        # Template donor tidak pernah diubah.
        try:
            from config import POKJA_ROOT as _pokja_root_header
            from document_profiles import apply_header_to_copy
            apply_header_to_copy(_merged_b, _pokja_root_header, data)
        except Exception as _header_err:
            raise RuntimeError(f"Gagal menerapkan header profil: {_header_err}")
        _blank_empty_participant_rows_xml(_merged_b, data)
        _strip_mailmerge_datasource(_merged_b)
        normalize_word_document_xml_in_zip(_merged_b)
        pythoncom.CoInitialize()
        wdApp = win32com.client.DispatchEx("Word.Application")
        wdApp.DisplayAlerts = 0
        wdApp.Visible = False
        try:
            wdDoc = wdApp.Documents.Open(
                FileName=_merged_b,
                ConfirmConversions=False,
                ReadOnly=False,
                AddToRecentFiles=False,
                Visible=False,
            )
            # re-merge field dari data Excel (satu_data) -> PDF selalu fresh
            if data:
                _replace_merge_fields(wdDoc, data)
                _trim_blank_participant_rows(wdDoc)
                _blank_empty_participant_rows(wdDoc, data)
                _protect_signature_layout(wdDoc)
                wdDoc.Save()
            if mode == "pdf_bapljkk":
                _kode_pljkk = pdf_name if pdf_name else "PL"
                _xlsm_path = None
                try:
                    import glob as _glob_pl
                    _xlsm_pl = _glob_pl.glob(os.path.join(_folder, "*.xlsm"))
                    if _xlsm_pl:
                        _xlsm_path = os.path.normpath(_xlsm_pl[0])
                        _xl_pl = win32com.client.DispatchEx("Excel.Application")
                        _xl_pl.Visible = False
                        _wb_pl = _xl_pl.Workbooks.Open(_xlsm_path, ReadOnly=True)
                        _ku_pl = str(_wb_pl.Sheets("@ Master Data").Range("G2").Value).strip()
                        _wb_pl.Close(False)
                        _xl_pl.Quit()
                        if _ku_pl and _ku_pl not in ("", "None", "null"):
                            _kode_pljkk = _ku_pl
                except Exception:
                    pass
                _pdf_path = _fit_path(_folder, f"BA_PLJKK_{_kode_pljkk}.pdf")
                _tmp_word = _pdf_path + "_tmpword.pdf"
                _tmp_72   = _pdf_path + "_tmp72.pdf"
                # Export Word -> tmp
                if wdDoc.Sections.Count >= 3:
                    _start = wdDoc.Sections(3).Range.Start
                else:
                    _start = wdDoc.Content.Start
                _rng = wdDoc.Range(_start, wdDoc.Content.End)
                _rng.ExportAsFixedFormat(OutputFileName=_tmp_word, ExportFormat=17)
                # Export sheet 7.2 Dengan Nego dari Excel -> tmp (jika ada)
                _has_72 = False
                if _xlsm_path:
                    try:
                        _xl72 = win32com.client.DispatchEx("Excel.Application")
                        _xl72.Visible = False
                        _wb72 = _xl72.Workbooks.Open(_xlsm_path, ReadOnly=True)
                        _ws72 = None
                        try:
                            _ws72 = _wb72.Sheets("7.2 Dengan Nego")
                        except Exception:
                            pass
                        if _ws72 is not None:
                            _ws72.ExportAsFixedFormat(
                                Type=0,  # xlTypePDF
                                Filename=_tmp_72,
                                Quality=0,
                                IncludeDocProperties=True,
                                IgnorePrintAreas=False,
                                OpenAfterPublish=False,
                            )
                            _has_72 = True
                        _wb72.Close(False)
                        _xl72.Quit()
                    except Exception:
                        try:
                            _xl72.Quit()
                        except Exception:
                            pass
                # Gabung PDF: word_part1 + sheet72(2x) + word_part2
                # Cari halaman "DAFTAR HADIR KLARIFIKASI DAN NEGOSIASI TEKNIS DAN BIAYA" di word PDF
                if _has_72:
                    try:
                        import pdfplumber
                        from pypdf import PdfWriter, PdfReader
                        _rdr_word = PdfReader(_tmp_word)
                        _n_word = len(_rdr_word.pages)
                        # Cari halaman pertama daftar hadir klarifikasi nego
                        _split_page = _n_word  # default: tidak ketemu = append di akhir
                        with pdfplumber.open(_tmp_word) as _plb:
                            for _pi, _pp in enumerate(_plb.pages):
                                _txt = (_pp.extract_text() or "").upper()
                                if "DAFTAR HADIR KLARIFIKASI DAN NEGOSIASI" in _txt:
                                    _split_page = _pi  # 0-based
                                    break
                        _rdr_72 = PdfReader(_tmp_72)
                        _writer = PdfWriter()
                        # Bagian 1: word halaman 0.._split_page-1
                        for _pi in range(_split_page):
                            _writer.add_page(_rdr_word.pages[_pi])
                        # Sheet 7.2: 2x
                        for _ in range(2):
                            for _pi in range(len(_rdr_72.pages)):
                                _writer.add_page(_rdr_72.pages[_pi])
                        # Bagian 2: word halaman _split_page..akhir
                        for _pi in range(_split_page, _n_word):
                            _writer.add_page(_rdr_word.pages[_pi])
                        with open(_pdf_path, "wb") as _fout:
                            _writer.write(_fout)
                    except Exception as _merge_err:
                        # Fallback: pakai word PDF saja
                        import shutil as _sh
                        _sh.copy2(_tmp_word, _pdf_path)
                    finally:
                        for _tf in (_tmp_word, _tmp_72):
                            try:
                                os.remove(_tf)
                            except Exception:
                                pass
                else:
                    # Tidak ada sheet 7.2, rename tmp -> final
                    import shutil as _sh
                    _sh.move(_tmp_word, _pdf_path)
                # Bentuk BA final sama seperti tombol "Gabung BA PLJKK".
                # File Summary SPSE tersimpan di subfolder 7, bukan di root,
                # sehingga helper lama tidak pernah menemukannya saat Cetak BA.
                try:
                    from gabung_ba_pljkk import gabung as _gabung_ba_pljkk
                    _gabung_result = _gabung_ba_pljkk(_folder)
                    if _gabung_result.get("ok"):
                        _pdf_path = _gabung_result["output"]
                except Exception:
                    pass
                show_success(_pdf_path)
            elif mode == "printer_bapljkk":
                # Printer harus memakai PDF final, bukan Word mentah.
                # Word mentah melewati Summary SPSE dan dua copy sheet 7.2.
                _printer_name = pdf_name
                _final_pdf = _build_bapljkk_final_pdf(wdDoc, _folder, "PL")
                import win32api
                _result = win32api.ShellExecute(
                    0, "printto", _final_pdf, f'"{_printer_name}"', _folder, 0
                )
                if _result <= 32:
                    raise RuntimeError(f"ShellExecute printto gagal ({_result})")
                time.sleep(3)
                show_print_success(_printer_name)
            wdDoc.Close(False)
        except Exception as e:
            show_error(f"Error cetak BA PLJKK:\n{e}")
        finally:
            wdApp.Quit()
            pythoncom.CoUninitialize()
            try:
                if os.path.exists(_merged_b):
                    import send2trash
                    send2trash.send2trash(_merged_b)
            except Exception:
                pass
        return

    folder = os.path.dirname(word_path)
    base, ext = os.path.splitext(os.path.basename(word_path))
    if ext.lower() not in (".docx", ".docm"):
        ext = ".docx"
    copy_path = _fit_path(folder, f"{base[:60].rstrip()} (Merged){ext}")

    # Copy template ke (Merged) - template asli tidak diubah
    shutil.copy2(word_path, copy_path)
    # Header resmi dipilih dari profil instansi, hanya pada copy sementara.
    try:
        from config import POKJA_ROOT as _pokja_root_header
        from document_profiles import apply_header_to_copy
        apply_header_to_copy(copy_path, _pokja_root_header, data)
    except Exception as _header_err:
        raise RuntimeError(f"Gagal menerapkan header profil: {_header_err}")
    _blank_empty_participant_rows_xml(copy_path, data)
    _strip_mailmerge_datasource(copy_path)

    pythoncom.CoInitialize()
    wdApp = None
    wdDoc = None
    word_pid = None
    new_instance = False
    _deferred_pdf_success = None

    try:
        wdApp = win32com.client.DispatchEx("Word.Application")
        new_instance = True

        wdApp.DisplayAlerts = 0
        wdApp.Visible = False

        _is_dokpil = (
            (data or {}).get("_source_sheet") == "list_dokpil"
            or os.path.basename(word_path).lower().startswith("3. dokpil full")
        )
        _is_reviu = (data or {}).get("_source_sheet") == "list_reviu"
        if _is_dokpil or _is_reviu:
            # Bentuk tabel sebelum Documents.Open: Word sering mengubah
            # struktur row/field ketika dokumen dibuka, terutama pada tabel
            # dengan vertically merged cells.
            _prepare_dokpil_equipment_docx(copy_path, data)
            _prepare_dokpil_personnel_docx(copy_path, data)

        # lxml-based preparation can collapse Word namespace declarations;
        # repair the copy immediately before COM parses it.
        normalize_word_document_xml_in_zip(copy_path)

        wdDoc = wdApp.Documents.Open(
            FileName=copy_path,
            ConfirmConversions=False,
            ReadOnly=False,
            AddToRecentFiles=False,
            Visible=False
        )
        word_pid = _word_process_id(wdApp, wdDoc)

        # Dokpil memiliki tabel personel dinamis. Bentuk row terlebih dahulu
        # saat field masih berupa MERGEFIELD; setelah di-unlink field clone
        # tidak lagi bisa diarahkan ke Personil 2/3.
        _replace_merge_fields(wdDoc, data)
        _trim_blank_participant_rows(wdDoc)
        _blank_empty_participant_rows(wdDoc, data)
        _protect_signature_layout(wdDoc)

        # Proteksi layout tanda tangan dapat mengubah pagination. Stabilkan
        # halaman terlebih dahulu, lalu tangani hanya field bookmark stale;
        # TOC asli tidak dibangun ulang agar entry/tab stop template tetap.
        try:
            wdDoc.Repaginate()
        except Exception:
            pass
        _update_toc_fields(wdDoc)

        # Cleanup blank pages untuk file BA utama (satu_data) yang multi-section.
        # File "2. Isi Reviu" & "3. Dokpil" dikecualikan (struktur beda, bisa berantakan).
        # Template BA dipecah per-dokumen: 4=Undangan, 5=BA Utama, 6=Ringkasan, 7=Timpang.
        _bn_cleanup = os.path.basename(word_path)
        if any(_bn_cleanup.startswith(_p) for _p in (
            "1. Full Dokumen", "4. Undangan", "5. Berita Acara", "6. Ringkasan", "7. BA Dengan"
        )):
            wdApp.ScreenUpdating = True
            if mode in ("buka", "print"):
                wdApp.Visible = True
                wdApp.WindowState = 2
            cleanup_blank_pages(wdDoc)

        # Simpan dan Tampilkan hanya jika bukan mode PDF
        if mode in ("buka", "print"):
            wdDoc.Save()
            wdApp.ScreenUpdating = True
            wdApp.Visible = True
            wdDoc.Windows(1).Visible = True
            wdDoc.Activate()
            wdDoc.Repaginate()
            time.sleep(1)

            try:
                import ctypes
                hwnd = ctypes.windll.user32.FindWindowW("OpusApp", None)
                if hwnd:
                    ctypes.windll.user32.ShowWindow(hwnd, 3)
                    ctypes.windll.user32.SetForegroundWindow(hwnd)
            except:
                wdApp.WindowState = 1

        if mode == "print":
            time.sleep(1)
            shell = win32com.client.Dispatch("WScript.Shell")
            shell.SendKeys("^p", 0)

        elif mode == "printer":
            # Direct print ke printer fisik (tanpa dialog)
            printer_name = pdf_name  # arg ke-5 dipakai sebagai nama printer
            from_page = int(sys.argv[6]) if len(sys.argv) > 6 else 0
            to_page = int(sys.argv[7]) if len(sys.argv) > 7 else 0

            if data.get("_source_sheet") == "satu_data":
                _strip_pl_ba_signature_header(wdDoc)

            wdDoc.Save()
            wdApp.ScreenUpdating = True
            wdApp.Visible = False

            try:
                # Set printer tujuan
                wdApp.ActivePrinter = printer_name

                # PrintOut dengan atau tanpa page range
                if from_page > 0 and to_page > 0:
                    wdDoc.PrintOut(
                        Background=False,
                        Range=3,   # wdPrintFromTo
                        From=str(from_page),
                        To=str(to_page),
                    )
                else:
                    wdDoc.PrintOut(Background=False)

                # Tunggu spooler selesai
                time.sleep(3)
                show_print_success(printer_name)
            except Exception as print_err:
                show_error(f"Gagal print ke {printer_name}:\n{print_err}")

            wdDoc.Close(False)
            if new_instance:
                wdApp.Quit()
            return  # skip cleanup di bawah

        elif mode == "printer_bapljkk":
            # Print Section 3 s/d akhir ke printer fisik (skip Reviu)
            printer_name = pdf_name  # arg ke-5 = nama printer

            wdDoc.Save()
            wdApp.ScreenUpdating = True
            wdApp.Visible = False

            try:
                wdApp.ActivePrinter = printer_name
                if wdDoc.Sections.Count >= 3:
                    _start = wdDoc.Sections(3).Range.Start
                else:
                    _start = wdDoc.Content.Start
                _rng = wdDoc.Range(_start, wdDoc.Content.End)
                _rng.Select()
                wdDoc.PrintOut(Background=False, Range=1)  # wdPrintSelection=1
                time.sleep(3)
                show_print_success(printer_name)
            except Exception as print_err:
                show_error(f"Gagal print ke {printer_name}:\n{print_err}")

            wdDoc.Close(False)
            if new_instance:
                wdApp.Quit()
            return

        elif mode.startswith("pdf"):
            # Ambil nama paket dari data dict
            _npk = ''
            for _nk in ['Nama_Paket','NamaTender','Nama_Tender','nama_paket','nama_tender']:
                _v = data.get(_nk) or data.get(_nk.lower())
                if _v and str(_v).strip() not in ('','None','null'):
                    _npk = str(_v).strip(); break
            # Tender Excel mengirim marker eksplisit agar nama PDF stabil
            # berdasarkan nomor Pokja, bukan nama paket yang panjang.
            nama_paket_pdf = _pdf_output_suffix(pdf_name, _npk)

            if mode == "pdf_full":
                # Export full dokumen (template BA dipecah per-file, no page-range).
                # Label dokumen dari nama file Word (prefix angka)
                _bn_full = os.path.basename(word_path)
                _label = "Dokumen"
                if _bn_full.startswith("1. BA Reviu") or _bn_full.startswith("1. Reviu"):
                    _label = "BA_REVIU_DPP"
                elif _bn_full.startswith("4. Undangan"):   _label = "Undangan"
                elif _bn_full.startswith("6. Ringkasan"):  _label = "REvaluasi"
                pdf_path = _fit_path(folder, f"{_label}_{nama_paket_pdf}.pdf")
                wdDoc.ExportAsFixedFormat(
                    OutputFileName=pdf_path,
                    ExportFormat=17,
                    Range=0,  # wdExportAllDocument
                )
                _deferred_pdf_success = pdf_path
            elif mode == "pdf_bareviu":
                pdf_path = _fit_path(folder, f"BA_REVIU_DPP_{nama_paket_pdf}.pdf")
                wdDoc.ExportAsFixedFormat(
                    OutputFileName=pdf_path,
                    ExportFormat=17,
                    Range=3,  # wdExportFromTo
                    From=3,
                    To=6,
                )
                _deferred_pdf_success = pdf_path
            elif mode == "pdf_bareviu_pl":
                _strip_pl_ba_signature_header(wdDoc)
                pdf_path = _fit_path(folder, f"BA_REVIU_PL_{nama_paket_pdf}.pdf")
                wdDoc.ExportAsFixedFormat(
                    OutputFileName=pdf_path,
                    ExportFormat=17,
                    Range=3,  # wdExportFromTo
                    From=1,
                    To=3,
                )
                _deferred_pdf_success = pdf_path
            elif mode == "pdf_bapljkk":
                # Export Section 3 s/d akhir (skip Section 1+2 = Reviu DPP)
                # Pakai Range agar tidak perlu tahu nomor halaman (robust terhadap perubahan isi Reviu)
                pdf_path = _fit_path(folder, f"BA_PLJKK_{nama_paket_pdf}.pdf")
                # File baru (BA-only) hanya 1 section; file lama (gabung Reviu) BA mulai Section 3
                if wdDoc.Sections.Count >= 3:
                    _start = wdDoc.Sections(3).Range.Start
                else:
                    _start = wdDoc.Content.Start
                _rng = wdDoc.Range(_start, wdDoc.Content.End)
                _rng.ExportAsFixedFormat(OutputFileName=pdf_path, ExportFormat=17)
                _deferred_pdf_success = pdf_path
            elif mode == "pdf_revaluasi":
                pdf_path = _fit_path(folder, f"REvaluasi_{nama_paket_pdf}.pdf")
                wdDoc.ExportAsFixedFormat(
                    OutputFileName=pdf_path,
                    ExportFormat=17,
                    Range=3,  # wdExportFromTo
                    From=30,
                    To=37,
                )
                _deferred_pdf_success = pdf_path
            elif mode == "pdf_all":
                # Output ke subfolder "6. BA Reviu Lengkap" (buat kalau belum ada)
                _ba_reviu_dir = os.path.join(folder, "6. BA Reviu Lengkap")
                os.makedirs(_ba_reviu_dir, exist_ok=True)
                pdf_path = _fit_path(_ba_reviu_dir, f"Isi_Reviu_DPP_{nama_paket_pdf}.pdf")
                try:
                    wdDoc.ExportAsFixedFormat(
                        OutputFileName=pdf_path,
                        ExportFormat=17,
                        Range=0,  # wdExportAllDocument
                    )
                except Exception:
                    wdDoc.SaveAs2(pdf_path, FileFormat=17)
                _deferred_pdf_success = pdf_path
            elif mode == "pdf_dokpil":
                # Ambil nama paket dari sheet satu_data (list_dokpil tidak punya field nama paket)
                _np_dokpil = nama_paket_pdf
                # `pdf_name` adalah marker mentah (mis. POKJA_041). Jika
                # suffix masih sama dengan marker/default mentah, ambil nama
                # paket dari satu_data sebagai fallback. Jangan pakai
                # `safe_name`: variabel itu hanya lokal di helper suffix.
                if _np_dokpil == (pdf_name or "000"):  # fallback belum dapat nama
                    try:
                        _data_sd = read_excel_data(excel_path, "satu_data")
                        if _data_sd:
                            for _nk in ['Nama_Paket', 'NamaTender', 'Nama_Tender', 'nama_paket']:
                                _v = _data_sd.get(_nk) or _data_sd.get(_nk.lower())
                                if _v and str(_v).strip() not in ('', 'None', 'null'):
                                    _np_dokpil = _safe_filename(str(_v).strip())
                                    break
                    except Exception:
                        pass
                pdf_path = _next_available_pdf_path(
                    _fit_path(folder, f"dokpil_{_np_dokpil}.pdf")
                )
                wdDoc.ExportAsFixedFormat(
                    OutputFileName=pdf_path,
                    ExportFormat=17,
                    Range=0,  # wdExportAllDocument
                )
                _deferred_pdf_success = pdf_path
            elif mode == "pdf_pembuktian":
                # File "5. Berita Acara Utama PK": export full Word -> PDF, sisip sheet
                # "7.2 Dengan Nego" SETELAH tiap halaman anchor nego (2 occurrence).
                # Anchor teks robust thd geseran halaman (ganti page-range/index manual).
                import tempfile
                final_pdf_path = _fit_path(folder, f"BA_Pembuktian_Nego_{nama_paket_pdf}.pdf")
                temp_dir = tempfile.mkdtemp()
                temp_word_pdf = os.path.join(temp_dir, "temp_word.pdf")
                temp_nego_pdf = os.path.join(temp_dir, "temp_nego.pdf")

                wdDoc.ExportAsFixedFormat(
                    OutputFileName=temp_word_pdf, ExportFormat=17, Range=0,
                )
                _has_nego = _export_sheet_pdf(excel_path, "7.2 Dengan Nego", temp_nego_pdf, landscape=True)

                _ANCHOR_NEGO = "DAFTAR HADIR NEGOSIASI KUANTITAS DAN HARGA"
                if _has_nego:
                    # 2 occurrence anchor -> sisip nego 2x (FIFO antrian di helper)
                    pairs = [(_ANCHOR_NEGO, temp_nego_pdf), (_ANCHOR_NEGO, temp_nego_pdf)]
                    final_pdf_path = _stitch_excel_at_anchor(temp_word_pdf, pairs, final_pdf_path)
                else:
                    import shutil as _sh
                    _sh.copy2(temp_word_pdf, final_pdf_path)
                _deferred_pdf_success = final_pdf_path

            elif mode == "pdf_pembuktian_timpang":
                # File "7. BA Dengan Timpang PK": export full Word -> PDF, sisip:
                #   - "7.2 Dengan Nego" setelah tiap anchor nego (2 occurrence)
                #   - "Klarifikasi Timpang Fix (2)" setelah tiap anchor timpang (2 occurrence)
                # Urutan sisip per halaman ditentukan posisi anchor di dokumen (robust).
                import tempfile
                final_pdf_path = _fit_path(folder, f"BA_Pembuktian_Timpang_{nama_paket_pdf}.pdf")
                temp_dir = tempfile.mkdtemp()
                temp_word_pdf = os.path.join(temp_dir, "temp_word.pdf")
                temp_nego_pdf = os.path.join(temp_dir, "temp_nego.pdf")
                temp_timpang_pdf = os.path.join(temp_dir, "temp_timpang.pdf")

                wdDoc.ExportAsFixedFormat(
                    OutputFileName=temp_word_pdf, ExportFormat=17, Range=0,
                )
                _has_nego = _export_sheet_pdf(excel_path, "7.2 Dengan Nego", temp_nego_pdf, landscape=True)
                _has_timpang = _export_sheet_pdf(
                    excel_path, "Klarifikasi Timpang Fix (2)", temp_timpang_pdf,
                    landscape=True, fit_wide=1, fit_tall=1,
                )

                _ANCHOR_NEGO = "DAFTAR HADIR NEGOSIASI KUANTITAS DAN HARGA"
                _ANCHOR_TIMPANG = "DAFTAR HADIR KLARIFIKASI HARGA SATUAN TIMPANG"
                pairs = []
                if _has_nego:
                    pairs += [(_ANCHOR_NEGO, temp_nego_pdf), (_ANCHOR_NEGO, temp_nego_pdf)]
                if _has_timpang:
                    pairs += [(_ANCHOR_TIMPANG, temp_timpang_pdf), (_ANCHOR_TIMPANG, temp_timpang_pdf)]

                if pairs:
                    final_pdf_path = _stitch_excel_at_anchor(temp_word_pdf, pairs, final_pdf_path)
                else:
                    import shutil as _sh
                    _sh.copy2(temp_word_pdf, final_pdf_path)
                _deferred_pdf_success = final_pdf_path

            else:
                pdf_path = _fit_path(folder, f"Undangan_{nama_paket_pdf}.pdf")
                wdDoc.ExportAsFixedFormat(
                    OutputFileName=pdf_path,
                    ExportFormat=17,
                    Range=3,  # wdExportFromTo
                    From=1,
                    To=2,
                )
                _deferred_pdf_success = pdf_path

            # Tutup Word dulu sebelum membuka PDF. COM cleanup sengaja
            # dipisahkan dan ditelan agar error "object disconnected" saat
            # Close/Quit tidak memunculkan popup atau membuka file Merged.
            try:
                wdDoc.Close(False)
            except Exception:
                pass
            if new_instance:
                try:
                    wdApp.Quit()
                except Exception:
                    pass

    except Exception as e:
        if wdApp:
            try:
                wdApp.ScreenUpdating = True
                wdApp.Visible = mode in ("buka", "print")
            except: pass
        show_error(f"Error saat merge:\n{e}")
    finally:
        if mode.startswith("pdf"):
            try:
                if wdDoc is not None:
                    wdDoc.Close(False)
            except Exception:
                pass
            try:
                if wdApp is not None and new_instance:
                    wdApp.Quit()
            except Exception:
                pass
            _terminate_word_process(word_pid)
        pythoncom.CoUninitialize()
        # Hapus file (Merged) ke Recycle Bin setelah selesai
        try:
            if os.path.exists(copy_path):
                import send2trash
                send2trash.send2trash(copy_path)
        except Exception:
            pass
        try:
            # Word meninggalkan lock file ``~$`` bila instance dihentikan
            # setelah COM terputus. File ini bukan dokumen user dan aman
            # dibersihkan setelah PID engine ditutup.
            lock_path = os.path.join(
                os.path.dirname(copy_path),
                "~$" + os.path.basename(copy_path)[2:],
            )
            if os.path.exists(lock_path):
                os.remove(lock_path)
        except Exception:
            pass
        # Buka PDF setelah Word dan file Merged benar-benar selesai ditutup.
        if _deferred_pdf_success:
            show_success(_deferred_pdf_success)


def show_error(msg):
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(0, msg, "Word Merge Error", 0x10)
    except:
        print(f"ERROR: {msg}")


def _safe_write_pdf(writer, target_path):
    """Write PDF, fallback ke suffix _v2/_v3/... jika file di-lock proses lain."""
    path = target_path
    for attempt in range(5):
        try:
            with open(path, 'wb') as f:
                writer.write(f)
            return path
        except PermissionError:
            base, ext = os.path.splitext(target_path)
            path = f"{base}_v{attempt + 2}{ext}"
    raise PermissionError(f"Gagal tulis PDF setelah 5 percobaan: {target_path}")


def show_success(pdf_path):
    """Notifikasi popup setelah PDF selesai dibuat, lalu buka file."""
    try:
        import ctypes
        filename = os.path.basename(pdf_path)
        ctypes.windll.user32.MessageBoxW(
            0, f"PDF berhasil dibuat:\n{filename}", "Export PDF Selesai", 0x40
        )
    except:
        pass
    try:
        if os.path.exists(pdf_path):
            os.startfile(pdf_path)
    except:
        pass


def show_print_success(printer_name):
    """Notifikasi popup setelah print dikirim ke spooler."""
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(
            0, f"Dokumen dikirim ke antrian print:\n{printer_name}\n\n"
               f"Pastikan printer menyala untuk mencetak.",
            "Print Dikirim", 0x40
        )
    except:
        pass


def run_merge_mode_pl(folder_path: str, excel_path: str) -> list:
    """
    Merge semua Word template PL di folder_path menggunakan data dari excel_path.
    Loop over WORD_SHEET_MAP_PL: (word_filename, sheet_name).
    Return: list hasil per file — {"file": str, "sukses": bool, "pesan": str}
    """
    import glob as _glob
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    try:
        from config import WORD_SHEET_MAP_PL
    except ImportError:
        WORD_SHEET_MAP_PL = [
            ("5. BA PLJKK - Template.docx",               "satu_data"),
            ("1. BA Reviu DPP PLJKK - Template.docx",      "list_reviu"),
            ("3. Dokpil Full PLJKK - Template.docx",      "list_dokpil"),
        ]

    results = []
    for word_filename, sheet_name in WORD_SHEET_MAP_PL:
        # Cari file — bisa Template atau sudah direname
        base = word_filename.replace(" - Template", "").replace(" - Template", "")
        candidates = _glob.glob(os.path.join(folder_path, word_filename))
        if not candidates:
            # Coba nama tanpa "- Template"
            stem = os.path.splitext(word_filename)[0].replace(" - Template", "").strip()
            ext  = os.path.splitext(word_filename)[1]
            candidates = _glob.glob(os.path.join(folder_path, f"{stem}*{ext}"))
        if not candidates:
            results.append({"file": word_filename, "sukses": False, "pesan": "File tidak ditemukan di folder"})
            continue

        word_path = candidates[0]
        data = read_excel_data(excel_path, sheet_name)
        if data is None:
            results.append({"file": os.path.basename(word_path), "sukses": False, "pesan": f"Gagal baca sheet {sheet_name}"})
            continue

        try:
            merge_word(word_path, data, mode="buka", pdf_name="")
            results.append({"file": os.path.basename(word_path), "sukses": True, "pesan": "OK"})
        except Exception as e:
            results.append({"file": os.path.basename(word_path), "sukses": False, "pesan": str(e)})

    return results


if __name__ == "__main__":
    if len(sys.argv) >= 2 and sys.argv[1].lower() == "personil":
        if len(sys.argv) != 5:
            print("Usage: python word_merge.py personil <template.docx> <excel.xlsm> <output.docx>")
            sys.exit(1)
        from merge_list_personil import merge_list_personil
        merge_list_personil(sys.argv[2], sys.argv[3], sys.argv[4])
        sys.exit(0)
    if len(sys.argv) < 5:
        print("Usage: python word_merge.py <mode> <word_path> <excel_path> <sheet_name>")
        print("  mode: buka | print | pdf | pdf_bareviu | pdf_bareviu_pl | personil")
        sys.exit(1)

    mode = sys.argv[1]
    word_path = sys.argv[2]
    excel_path = sys.argv[3]
    sheet_name = sys.argv[4]
    pdf_name = sys.argv[5] if len(sys.argv) > 5 else ""

    data = read_excel_data(excel_path, sheet_name)

    if data is None:
        sys.exit(1)

    merge_word(word_path, data, mode, pdf_name)
