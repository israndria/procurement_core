"""
setup_master_data.py — Setup sheet "@ Master Data" di workbook .xlsm.

Fungsi:
  1. Buat sheet "@ Master Data" (atau rename "_HasilParse" jika ada)
  2. Tulis header + section label di baris tetap
  3. Tulis formula di sheet target: "1. Input Data", "database_reviu", "database_dokpil"
     → kolom E merujuk ke "@ Master Data" kolom C

Dijalankan via:
  - Langsung: python setup_master_data.py <path_to_xlsm>
  - Dari inject_buttons.py: dipanggil setelah inject VBA
"""

import sys
import os

# Coba openpyxl dulu (lebih cepat, tidak butuh Excel terbuka)
try:
    import openpyxl
    USE_OPENPYXL = True
except ImportError:
    USE_OPENPYXL = False

# ── Layout @ Master Data: baris tetap per field ─────────────────────────────
MASTER_SHEET = "@ Master Data"

# Row 1 = header, Row 2 = section header INPUT DATA
INPUT_DATA_MAP = {
    # cell di "1. Input Data" → baris di @ Master Data
    "E3":  3,   # Kode Rekening (MAK)
    "E5":  4,   # Kode Tender
    "E6":  5,   # Nama Tender
    "E8":  6,   # Kode RUP
    "E10": 7,   # Nilai Pagu
    "E11": 8,   # Nilai HPS
    "E12": 9,   # Nomor Surat Permohonan
    "E13": 10,  # Nomor Surat Tugas
    "E14": 11,  # Kode Pokja
    "E15": 12,  # Masa Pelaksanaan (Hari)
    "E16": 13,  # Kegiatan/Sub Kegiatan (dari PDF)
    "E17": 14,  # SKPD/OPD
    "E19": 15,  # Nama PPK
    "E20": 16,  # NIP PPK
    "E21": 17,  # Nomor SK PPK
    "E22": 18,  # Anggota 1
    "E23": 19,  # Anggota 2
    "E24": 20,  # Anggota 3
    "E32": 21,  # Lokasi (dari PDF)
    "E33": 22,  # Sumber Dana
}

INPUT_DATA_LABELS = {
    "E3": "Kode Rekening (MAK)",
    "E5": "Kode Tender",
    "E6": "Nama Tender",
    "E8": "Kode RUP",
    "E10": "Nilai Pagu",
    "E11": "Nilai HPS",
    "E12": "Nomor Surat Permohonan",
    "E13": "Nomor Surat Tugas",
    "E14": "Kode Pokja",
    "E15": "Masa Pelaksanaan (Hari)",
    "E16": "Kegiatan/Sub Kegiatan",
    "E17": "SKPD/OPD",
    "E19": "Nama PPK",
    "E20": "NIP PPK",
    "E21": "Nomor SK PPK",
    "E22": "Anggota 1",
    "E23": "Anggota 2",
    "E24": "Anggota 3",
    "E32": "Lokasi",
    "E33": "Sumber Dana",
}

# Row 23 = kosong, Row 24 = section header DATABASE REVIU
REVIU_MAP = {
    "E2":  25,
    "E6":  26,  "E7":  27,
    "E9":  28,  "E10": 29,  "E11": 30,
    "E12": 31,  "E13": 32,  "E14": 33,
    "E15": 34,  "E16": 35,  "E17": 36,
    "E18": 37,  "E19": 38,  "E20": 39,
    "E21": 40,  "E22": 41,  "E23": 42,
    "E24": 43,  "E25": 44,  "E26": 45,
    "E27": 46,  "E28": 47,  "E29": 48,
    "E30": 49,  "E31": 50,
    "E32": 51,  "E33": 52,  "E34": 53,
}

REVIU_LABELS = {
    "E2": "Fungsi Bangunan",
    "E6": "Peralatan Utama", "E7": "Personil Manajerial",
    "E9": "Rencana K3K", "E10": "Analisis Harga Satuan",
    "E11": "Jenis Kontrak", "E12": "Bentuk Kontrak",
    "E13": "Masa Pertanggungan", "E14": "Jenis Penilaian",
    "E15": "Ambang Batas", "E16": "Jaminan Penawaran",
    "E17": "Jaminan Sanggah", "E18": "Jaminan Pelaksanaan",
    "E19": "Persyaratan SBU", "E20": "SBU Subkualifikasi",
    "E21": "SBU Kualifikasi", "E22": "Pengalaman Perusahaan",
    "E23": "Tenaga Ahli", "E24": "Peralatan",
    "E25": "Masa Berlaku Penawaran", "E26": "Jangka Waktu",
    "E27": "Jadwal Pelaksanaan", "E28": "Izin Usaha",
    "E29": "Kinerja", "E30": "NPWP", "E31": "SKP",
    "E32": "Uraian Pekerjaan", "E33": "Uraian Lingkup",
    "E34": "SBU Klasifikasi",
}

# Row 54 = kosong, Row 55 = section header DATABASE DOKPIL
DOKPIL_MAP = {
    "E6":  56,  "E7":  57,  "E8":  58,
    "E9":  59,  "E10": 60,  "E11": 61,
    "E12": 62,  "E13": 63,  "E14": 64,
    "E15": 65,  "E16": 66,
}

DOKPIL_LABELS = {
    "E6": "Syarat Kualifikasi 1", "E7": "Syarat Kualifikasi 2",
    "E8": "Syarat Kualifikasi 3", "E9": "Syarat Kualifikasi 4",
    "E10": "Syarat Kualifikasi 5", "E11": "Syarat Kualifikasi 6",
    "E12": "Syarat Kualifikasi 7", "E13": "Syarat Kualifikasi 8",
    "E14": "Syarat Kualifikasi 9", "E15": "Syarat Kualifikasi 10",
    "E16": "Cara Pembayaran",
}


def setup_with_openpyxl(wb_path):
    """Setup @ Master Data menggunakan openpyxl (tanpa Excel terbuka)."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(wb_path, keep_vba=True)

    # Rename _HasilParse jika ada, atau buat baru
    if "_HasilParse" in wb.sheetnames:
        ws = wb["_HasilParse"]
        ws.title = MASTER_SHEET
    elif MASTER_SHEET in wb.sheetnames:
        ws = wb[MASTER_SHEET]
    else:
        ws = wb.create_sheet(MASTER_SHEET, 0)

    # Pindah ke posisi pertama
    wb.move_sheet(ws, offset=-wb.sheetnames.index(MASTER_SHEET))

    # Clear isi
    for row in ws.iter_rows(min_row=1, max_row=70, max_col=4):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()

    # Header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(["Field", "Status", "Nilai", "Navigasi"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Section headers
    section_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    section_font = Font(bold=True)
    for row, title in [(2, "1. INPUT DATA"), (24, "DATABASE REVIU"), (55, "DATABASE DOKPIL")]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value=title)
        cell.fill = section_fill
        cell.font = section_font

    # Write labels di kolom A
    for cell_addr, row_num in INPUT_DATA_MAP.items():
        ws.cell(row=row_num, column=1, value=f"{INPUT_DATA_LABELS.get(cell_addr, cell_addr)} ({cell_addr})")
    for cell_addr, row_num in REVIU_MAP.items():
        ws.cell(row=row_num, column=1, value=f"{REVIU_LABELS.get(cell_addr, cell_addr)} ({cell_addr})")
    for cell_addr, row_num in DOKPIL_MAP.items():
        ws.cell(row=row_num, column=1, value=f"{DOKPIL_LABELS.get(cell_addr, cell_addr)} ({cell_addr})")

    # Kolom C = tempat user edit (kosong dulu, akan diisi VBA saat pilih paket)
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 15

    # ── Setup formula di sheet target ────────────────────────────────────────
    master_ref = f"'{MASTER_SHEET}'"

    # "1. Input Data"
    if "1. Input Data" in wb.sheetnames:
        ws_inp = wb["1. Input Data"]
        for cell_addr, master_row in INPUT_DATA_MAP.items():
            cell = ws_inp[cell_addr]
            cell.value = f"={master_ref}!C{master_row}"
        print(f"  [OK] 1. Input Data: {len(INPUT_DATA_MAP)} formula ditulis")
    else:
        print("  [SKIP] Sheet '1. Input Data' tidak ditemukan")

    # "database_reviu"
    if "database_reviu" in wb.sheetnames:
        ws_rev = wb["database_reviu"]
        for cell_addr, master_row in REVIU_MAP.items():
            cell = ws_rev[cell_addr]
            cell.value = f"={master_ref}!C{master_row}"
        print(f"  [OK] database_reviu: {len(REVIU_MAP)} formula ditulis")
    else:
        print("  [SKIP] Sheet 'database_reviu' tidak ditemukan")

    # "database_dokpil"
    if "database_dokpil" in wb.sheetnames:
        ws_dok = wb["database_dokpil"]
        for cell_addr, master_row in DOKPIL_MAP.items():
            cell = ws_dok[cell_addr]
            cell.value = f"={master_ref}!C{master_row}"
        print(f"  [OK] database_dokpil: {len(DOKPIL_MAP)} formula ditulis")
    else:
        print("  [SKIP] Sheet 'database_dokpil' tidak ditemukan")

    wb.save(wb_path)
    print(f"  [SAVED] {wb_path}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python setup_master_data.py <path_to_xlsm>")
        print("       python setup_master_data.py --all   (scan semua .xlsm di POKJA_ROOT)")
        sys.exit(1)

    arg = sys.argv[1]

    if arg == "--all":
        pokja_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        # Scan semua subfolder paket
        count = 0
        for root, dirs, files in os.walk(pokja_root):
            # Skip non-paket folders
            if any(skip in root for skip in [".git", "__pycache__", "node_modules", "V19_Scheduler", "Asisten_Pokja"]):
                continue
            for f in files:
                if f.endswith(".xlsm") and "BA PK" in f and "Improved" in f:
                    fp = os.path.join(root, f)
                    print(f"\n--- {f} ---")
                    try:
                        setup_with_openpyxl(fp)
                        count += 1
                    except Exception as e:
                        print(f"  [ERROR] {e}")
        print(f"\n=== {count} workbook(s) processed ===")
    else:
        if not os.path.exists(arg):
            print(f"File tidak ditemukan: {arg}")
            sys.exit(1)
        setup_with_openpyxl(arg)


if __name__ == "__main__":
    main()
