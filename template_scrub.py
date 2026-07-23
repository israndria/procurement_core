"""Scrub paket baru dari data donor tanpa merusak layout/macro."""

from __future__ import annotations

import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
W = f"{{{W_NS}}}"


def _cell_text(value) -> str:
    return str(value).strip() if value not in (None, "") else ""


def _excel_markers(excel_path: Path) -> list[str]:
    """Ambil marker donor dari workbook tanpa menyimpan ulang workbook."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(excel_path), read_only=True, data_only=True, keep_vba=True)
        markers: list[str] = []
        for sheet_name in ("@ Master Data", "0. Input BA", "3. KK Evaluasi Kualifikasi"):
            if sheet_name not in wb.sheetnames:
                continue
            for row in wb[sheet_name].iter_rows():
                for cell in row:
                    value = _cell_text(cell.value)
                    if value and any(x in value.lower() for x in (
                        "kode unik", "kode tender", "nama tender", "kode pokja",
                        "pembangunan", "normalisasi", "pengerukkan", "cv ",
                        "pt ", "firma ", "rup", "pokja",
                    )):
                        markers.append(value)
        wb.close()
        return sorted(set(markers), key=len, reverse=True)
    except Exception:
        return []


def _clear_docm_content_controls(data: bytes) -> bytes:
    """Kosongkan isi CC, pertahankan tag, format, dan struktur CC."""
    with tempfile.TemporaryDirectory() as td:
        src, out = Path(td) / "in.docm", Path(td) / "out.docm"
        src.write_bytes(data)
        with zipfile.ZipFile(src, "r") as zin:
            if "word/document.xml" not in zin.namelist():
                return data
            root = ET.fromstring(zin.read("word/document.xml"))
            count = 0
            for sdt in root.findall(f".//{W}sdt"):
                content = sdt.find(f"{W}sdtContent")
                if content is None:
                    continue
                for node in content.iter():
                    if node.tag == f"{W}t":
                        node.text = ""
                count += 1
            if not count:
                return data
            new_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    payload = new_xml if item.filename == "word/document.xml" else zin.read(item.filename)
                    zout.writestr(item, payload)
            return out.read_bytes()


def _replace_markers_in_office(data: bytes, markers: list[str]) -> bytes:
    """Hapus marker donor dari XML Word; format dokumen tetap."""
    if not markers:
        return data
    with tempfile.TemporaryDirectory() as td:
        src, out = Path(td) / "in.docx", Path(td) / "out.docx"
        src.write_bytes(data)
        changed = False
        with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                payload = zin.read(item.filename)
                if item.filename.startswith("word/") and item.filename.endswith(".xml"):
                    text = payload.decode("utf-8", "ignore")
                    new_text = text
                    for marker in markers:
                        new_text = new_text.replace(marker, "")
                    if new_text != text:
                        payload = new_text.encode("utf-8")
                        changed = True
                zout.writestr(item, payload)
        return out.read_bytes() if changed else data


def _clear_constants_com(excel_path: Path, targets: dict[str, list[str]]) -> list[str]:
    """Clear konstanta via satu sesi Excel COM. Formula/macro/shape tetap utuh."""
    import pythoncom
    import win32com.client as win32

    pythoncom.CoInitialize()
    xl = win32.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        try:
            xl.AutomationSecurity = 1
        except Exception:
            pass
        wb = xl.Workbooks.Open(str(excel_path), UpdateLinks=0, ReadOnly=False)
        logs: list[str] = []
        # xlCellTypeConstants = 2; satu ClearContents jauh lebih cepat daripada
        # iterasi COM per sel, terutama sheet KK yang lebarnya 195 kolom.
        for sheet_name, ranges in targets.items():
            ws = wb.Worksheets(sheet_name)
            cleared = 0
            for address in ranges:
                try:
                    constants = ws.Range(address).SpecialCells(2)
                    constants.ClearContents()
                    cleared += 1
                except Exception:
                    # 1004 = tidak ada konstanta pada area; aman diabaikan.
                    pass
            logs.append(f"Excel scrub {sheet_name}: {cleared} area dibersihkan")
        wb.Save()
        wb.Close(SaveChanges=False)
        return logs
    finally:
        xl.Quit()
        pythoncom.CoUninitialize()


def scrub_excel_copy(excel_path: str | Path) -> list[str]:
    """Bersihkan area data donor dari copy workbook tender baru."""
    path = Path(excel_path)
    logs: list[str] = []
    targets = {
        "@ Master Data": ["C3:C70", "H2:I23"],
        "0. Input BA": ["C3:G5", "C7:E14", "G7:G14", "C17:E22", "G17:N22", "C25:C29", "C32:G33"],
        "3. KK Evaluasi Kualifikasi": ["C3:GR92"],
        "6. Harga Penawaran": ["A1:Z171"],
        "6. Harga Penawaran (2)": ["A1:Z171"],
        "6. Harga Penawaran (3)": ["A1:Z171"],
        "5. HPS": ["A1:Z300"],
    }
    try:
        logs.extend(_clear_constants_com(path, targets))
    except Exception as exc:
        logs.append(f"⚠ Excel scrub gagal: {exc}")
    return logs


def scrub_package_copy(target_dir: str | Path, excel_path: str | Path,
                       word_paths: list[str | Path]) -> list[str]:
    """Scrub copy baru + hasilkan log audit."""
    target = Path(target_dir)
    excel = Path(excel_path)
    markers = _excel_markers(excel)
    logs = scrub_excel_copy(excel)
    for word in word_paths:
        path = Path(word)
        if not path.exists() or path.suffix.lower() not in (".docx", ".docm"):
            continue
        try:
            original = path.read_bytes()
            data = _replace_markers_in_office(original, markers)
            if path.suffix.lower() == ".docm":
                data = _clear_docm_content_controls(data)
            if data != original:
                path.write_bytes(data)
                logs.append(f"Word scrub: {path.name}")
        except Exception as exc:
            logs.append(f"⚠ Word scrub {path.name} gagal: {exc}")

    for name in ("jawaban_reviu.json", "_parse_reviu.json"):
        stale = target / name
        if stale.exists():
            stale.unlink()
            logs.append(f"Hapus state donor: {name}")
    return logs
