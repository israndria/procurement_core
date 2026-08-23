"""Build the tender BA Minimalis Word template from a filled donor document.

The donor is a package-specific document.  This one-off builder keeps its
layout, then restores the Excel merge fields needed by ``word_merge.py``.
It never edits the donor in place.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pythoncom
import win32com.client
from openpyxl import load_workbook


WD_FIND_STOP = 0
WD_COLLAPSE_END = 0
WD_FIELD_MERGE = 59


def _cell_text(cell) -> str:
    return str(cell.Range.Text).rstrip("\r\a").strip()


def _field_at(doc, start: int, name: str) -> None:
    point = doc.Range(start, start)
    doc.Fields.Add(point, WD_FIELD_MERGE, name, True)


def replace_long_paragraphs(doc, old: str, field_name: str) -> int:
    count = 0
    for paragraph in doc.Paragraphs:
        visible = str(paragraph.Range.Text).rstrip("\r\a")
        if visible != old:
            continue
        rng = paragraph.Range.Duplicate
        rng.End = max(rng.Start, rng.End - 1)
        rng.Text = ""
        doc.Fields.Add(rng, WD_FIELD_MERGE, field_name, True)
        count += 1
    return count


def replace_all(doc, old: str, field_name: str) -> int:
    """Replace all exact visible occurrences with one MERGEFIELD."""
    if not old or old == "0":
        return 0
    # Word's Find.Text property rejects strings longer than 255 characters.
    if len(old) > 255:
        return replace_long_paragraphs(doc, old, field_name)
    count = 0
    cursor = doc.Content.Start
    while cursor < doc.Content.End:
        rng = doc.Range(cursor, doc.Content.End)
        find = rng.Find
        find.ClearFormatting()
        find.Replacement.ClearFormatting()
        find.Text = old
        find.Forward = True
        find.Wrap = WD_FIND_STOP
        find.Format = False
        if not find.Execute():
            break
        start = rng.Start
        doc.Fields.Add(rng, WD_FIELD_MERGE, field_name, True)
        count += 1
        cursor = max(start + 1, rng.End)
    return count


def replace_nth(doc, old: str, field_name: str, occurrence: int) -> bool:
    """Replace one zero-based occurrence, preserving date context."""
    seen = 0
    cursor = doc.Content.Start
    while cursor < doc.Content.End:
        rng = doc.Range(cursor, doc.Content.End)
        find = rng.Find
        find.ClearFormatting()
        find.Text = old
        find.Forward = True
        find.Wrap = WD_FIND_STOP
        find.Format = False
        if not find.Execute():
            return False
        if seen == occurrence:
            start = rng.Start
            doc.Fields.Add(rng, WD_FIELD_MERGE, field_name, True)
            return True
        seen += 1
        cursor = max(rng.End, cursor + 1)
    return False


def replace_all_with_prefix(doc, old: str, prefix: str, field_name: str) -> int:
    count = 0
    cursor = doc.Content.Start
    while cursor < doc.Content.End:
        rng = doc.Range(cursor, doc.Content.End)
        find = rng.Find
        find.ClearFormatting()
        find.Text = old
        find.Forward = True
        find.Wrap = WD_FIND_STOP
        find.Format = False
        if not find.Execute():
            break
        rng.Text = prefix
        rng.Collapse(WD_COLLAPSE_END)
        doc.Fields.Add(rng, WD_FIELD_MERGE, field_name, True)
        count += 1
        cursor = max(rng.End, cursor + 1)
    return count


def replace_cell(cell, field_name: str) -> None:
    rng = cell.Range.Duplicate
    rng.End = max(rng.Start, rng.End - 1)  # keep end-of-cell marker
    rng.Text = ""
    doc = cell.Range.Document
    doc.Fields.Add(rng, WD_FIELD_MERGE, field_name, True)


def restore_participant_cells(doc) -> int:
    """Restore participant/director merge fields in repeated signature tables."""
    changed = 0
    for table in doc.Tables:
        if "Peserta Tender" not in table.Range.Text or "Direktur" not in table.Range.Text:
            continue
        zeros = []
        try:
            for cell in table.Range.Cells:
                if _cell_text(cell) == "0":
                    zeros.append(cell)
        except Exception:
            continue
        split = max(1, len(zeros) // 2)
        for index, cell in enumerate(zeros):
            try:
                replace_cell(cell, "Peserta_1" if index < split else "Dirut_Peserta_1")
                changed += 1
            except Exception:
                pass
    return changed


def restore_participant_detail_table(doc) -> int:
    changed = 0
    for table in doc.Tables:
        if "Nama Peserta" not in table.Range.Text or "Nomor NPWP" not in table.Range.Text:
            continue
        for row in table.Rows:
            try:
                label = _cell_text(row.Cells(1)).casefold()
                if label.startswith("nama peserta"):
                    replace_cell(row.Cells(row.Cells.Count), "Peserta_1")
                    changed += 1
                elif label.startswith("nomor npwp"):
                    replace_cell(row.Cells(row.Cells.Count), "Nomor_NPWP")
                    changed += 1
                elif label.startswith("alamat kantor"):
                    replace_cell(row.Cells(row.Cells.Count), "Alamat_Kantor")
                    changed += 1
            except Exception:
                pass
    return changed


def build(source: Path, destination: Path, excel_path: Path) -> None:
    if destination.exists():
        raise FileExistsError(f"Destination already exists: {destination}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)

    wb = load_workbook(excel_path, read_only=True, data_only=True, keep_links=False)
    ws = wb["satu_data"]
    values = {str(h.value).strip(): v.value for h, v in zip(ws[1], ws[2]) if h.value}
    wb.close()

    app = win32com.client.DispatchEx("Word.Application")
    app.Visible = False
    app.DisplayAlerts = 0
    doc = None
    try:
        doc = app.Documents.Open(
            FileName=str(destination),
            ConfirmConversions=False,
            ReadOnly=False,
            AddToRecentFiles=False,
            Visible=False,
        )

        # Whole narrative/number fields must be restored before shorter fields.
        long_fields = [
            ("No BA Klarif Alat", "No_BA_Klarif_Alat"),
            ("Pembukaan BA Klarif Alat", "Pembukaan_BA_Klarif_Alat"),
            ("No BA Klarif", "No_BA_Klarif"),
            ("Pembukaan BA Klarif", "Pembukaan_BA_Klarif"),
            ("Isi BA Klarif 2", "Isi_BA_Klarif_2"),
            ("Isi BA Klarif", "Isi_BA_Klarif"),
            ("No BA Penetapan", "No_BA_Penetapan"),
        ]
        for header, field_name in long_fields:
            value = values.get(header)
            if value:
                replace_all(doc, str(value), field_name)

        # The donor uses the same date on the first three attendance blocks,
        # while the final attendance block belongs to penetapan.
        date_rantau = values.get("Tanggal Pembuktian")
        if date_rantau:
            for occurrence in (0, 1, 2):
                replace_nth(doc, str(date_rantau), "Tanggal_Pembuktian", occurrence)
            replace_nth(doc, str(date_rantau), "Tanggal_Penetapan", 3)
        date_full = values.get("Tanggal Pembuktian Lengkap")
        if date_full:
            replace_all(doc, str(date_full), "Tanggal_Pembuktian_Lengkap")

        # Metadata and final-price fields occur in table cells and paragraphs.
        fields = [
            ("Kode Tender", "Kode_Tender"),
            ("Nama Paket", "Nama_Paket"),
            ("SKPD/OPD", "SKPDOPD"),
            ("Sumber Pendanaan", "Sumber_Pendanaan"),
            ("Pagu", "Pagu"),
            ("HPS", "HPS"),
            ("Metode Kualifikasi", "Metode_Kualifikasi"),
            ("Metode Evaluasi Penawaran", "Metode_Evaluasi_Penawaran"),
            ("SubKegiatan", "SubKegiatan"),
            ("Lokasi", "Lokasi"),
            ("Harga Terkoreksi", "Harga_Terkoreksi"),
            ("Harga Penawaran", "Harga_Penawaran"),
            ("Harga Nego", "Harga_Nego"),
            ("Waktu Pelaksanaan", "Waktu_Pelaksanaan"),
            ("Nomor NPWP", "Nomor_NPWP"),
        ]
        for header, field_name in fields:
            value = values.get(header)
            if value not in (None, "", 0, "0"):
                replace_all(doc, str(value), field_name)

        replace_all_with_prefix(
            doc,
            "Kelompok Kerja Pemilihan 057",
            "Kelompok Kerja Pemilihan ",
            "Pokja",
        )

        # Tender is a common label in the document, so replace only exact
        # table cells instead of every visible occurrence.
        method = values.get("Metode Pemilihan")
        if method:
            for table in doc.Tables:
                try:
                    for cell in table.Range.Cells:
                        if _cell_text(cell) == str(method):
                            replace_cell(cell, "Metode_Pemilihan")
                except Exception:
                    pass

        # Final price cell includes the numeric amount plus its terbilang.
        rounded = str(values.get("Harga Pembulatan") or "")
        words = str(values.get("Terbilang Pembulatan") or "")
        if rounded and words:
            replace_all(doc, f"{rounded} {words}", "Harga_Pembulatan")

        restore_participant_cells(doc)
        restore_participant_detail_table(doc)
        doc.Save()
        print(f"[OK] template: {destination}")
        print(f"[OK] merge fields: {doc.Fields.Count}")
    finally:
        if doc is not None:
            doc.Close(False)
        app.Quit()


if __name__ == "__main__":
    if len(sys.argv) != 4:
        raise SystemExit("usage: create_ba_minimalis_template.py SOURCE DESTINATION EXCEL")
    pythoncom.CoInitialize()
    try:
        build(Path(sys.argv[1]), Path(sys.argv[2]), Path(sys.argv[3]))
    finally:
        pythoncom.CoUninitialize()
